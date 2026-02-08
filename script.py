import streamlit as st
import osmnx as ox
import re
import geopandas as gpd
from openai import OpenAI
from shapely.geometry import Polygon, box
from selenium.webdriver.support.ui import WebDriverWait
from concurrent.futures import ThreadPoolExecutor, as_completed
import folium
from streamlit_folium import st_folium
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border,Side
from geopy.geocoders import Nominatim
import ssl
import zipfile
import json
import certifi
import concurrent.futures
import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
import shutil
import gzip
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from google_news import GoogleNews
import openai
from selenium.webdriver.common.by import By
from fuzzywuzzy import process, fuzz
from urllib.parse import urlparse
from tqdm import tqdm
import random
import base64
import mimetypes


def get_certificate_pem(hostname, port=443, cert_file='server_cert.pem'):
    cert = ssl.get_server_certificate((hostname, port))
    with open(cert_file, 'w') as f:
        f.write(cert)
    return cert_file

def extract_hostname(url):
    return urlparse(url).hostname

def combine_cert_with_certifi(cert_file, combined_file='combined.pem'):
    with open(certifi.where(), 'rb') as base, open(cert_file, 'rb') as site:
        combined = base.read() + b'\n' + site.read()
    with open(combined_file, 'wb') as f:
        f.write(combined)
    return combined_file

def secure_request_with_cert(url, combined_cert_file):
    response = requests.get(url, verify=combined_cert_file)
    # print(f"✅ Status: {response.status_code}")
    return response


route = st.query_params.get("page", "home")


css="""
    <style> 
    .stAppDeployButton {
            visibility: hidden;s
        } 
.st-emotion-cache-16tyu1 h2
 {
    font-size: 2.25rem;
    padding: 1rem 0px;
}
.st-emotion-cache-zy6yx3 {
    width: 100%;
    padding: 0.3 1.25em;
    max-width: 68em;
    margin: 0 auto;
}

}

.st-emotion-cache-16tyu1 h1{
font-size:  3.5rem;
color:#585858;
font-family: "Source Sans Pro", Helvetica, sans-serif;
}

.news_Consolidation{
font-size:  3.5rem;
color:#585858;
font-family: "Source Sans Pro", Helvetica, sans-serif;
}



html, body, div, span, applet, object,
iframe, h1, h2, h3, h4, h5, h6, p, blockquote,
pre, a, abbr, acronym, address, big, cite,
code, del, dfn, em, img, ins, kbd, q, s, samp,
small, strike, strong, sub, sup, tt, var, b,
u, i, center, dl, dt, dd, ol, ul, li, fieldset,
form, label, legend, table, caption, tbody,
tfoot, thead, tr, th, td, article, aside,
canvas, details, embed, figure, figcaption,
footer, header, hgroup, menu, nav, output, ruby,
section, summary, time, mark, audio, video {
	margin: 0;
	padding: 0;
	border: 0;
	font-size: 100%;
	font: inherit;
    font-family: "Source Sans Pro", Helvetica, sans-serif;
	vertical-align: baseline;}

article, aside, details, figcaption, figure,
footer, header, hgroup, menu, nav, section {
	display: block;}

body {
	line-height: 1;
}

ol, ul {
	list-style: none;
}

blockquote, q {
	quotes: none;
}

	blockquote:before, blockquote:after, q:before, q:after {
		content: '';
		content: none;
	}


table {
	border-collapse: collapse;
	border-spacing: 0;
}

body {
	-webkit-text-size-adjust: none;
}

mark {
	background-color: transparent;
	color: inherit;
}

input::-moz-focus-inner {
	border: 0;
	padding: 0;
}

input, select, textarea {
	-moz-appearance: none;
	-webkit-appearance: none;
	-ms-appearance: none;
	appearance: none;
}

/* Basic */

	@-ms-viewport {
		width: device-width;
	}

	body {
		-ms-overflow-style: scrollbar;
	}

	@media screen and (max-width: 480px) {

		html, body {
			min-width: 320px;
		}

	}

	html {
		box-sizing: border-box;
	}

	*, *:before, *:after {
		box-sizing: inherit;
	}

	body {
		background: #ffffff;
	}

		body.is-preload *, body.is-preload *:before, body.is-preload *:after {
			-moz-animation: none !important;
			-webkit-animation: none !important;
			-ms-animation: none !important;
			animation: none !important;
			-moz-transition: none !important;
			-webkit-transition: none !important;
			-ms-transition: none !important;
			transition: none !important;
		}

/* Type */

	body, input, select, textarea {
		color: #585858;
		font-family: "Source Sans Pro", Helvetica, sans-serif;
		font-size: 16pt;
		font-weight: 300;
		line-height: 1.75;
	}

		@media screen and (max-width: 1680px) {

			body, input, select, textarea {
				font-size: 14pt;
			}

		}

		@media screen and (max-width: 1280px) {

			body, input, select, textarea {
				font-size: 12pt;
			}

		}

	a {
		-moz-transition: border-bottom-color 0.2s ease, color 0.2s ease;
		-webkit-transition: border-bottom-color 0.2s ease, color 0.2s ease;
		-ms-transition: border-bottom-color 0.2s ease, color 0.2s ease;
		transition: border-bottom-color 0.2s ease, color 0.2s ease;
		text-decoration: none;
		color: #585858;
		border-bottom: dotted 1px rgba(88, 88, 88, 0.5);
	}

		a:hover {
			border-bottom-color: transparent;
			color: #f2849e !important;
		}

	strong, b {
		font-weight: 900;
	}

	em, i {
		font-style: italic;
	}

	p {
		margin: 0 0 2em 0;
	}

	h1 {
		font-size: 2.75em;
		font-weight: 00;
		line-height: 1.3;
		margin: 0 0 1em 0;
		letter-spacing: -0.035em;
	}

		h1 a {
			color: inherit;
		}

		@media screen and (max-width: 736px) {

			h1 {
				font-size: 2em;
				margin: 0 0 1em 0;
			}

		}

		@media screen and (max-width: 360px) {

			h1 {
				font-size: 1.75em;
			}

		}

	h2, h3, h4, h5, h6 {
		font-weight: 900;
		line-height: 1.5;
		margin: 0 0 2em 0;
		text-transform: uppercase;
		letter-spacing: 0.35em;
	}

		h2 a, h3 a, h4 a, h5 a, h6 a {
			color: inherit;
		}

	h2 {
		font-size: 1.1em;
	}

	h3 {
		font-size: 1em;
	}

	h4 {
		font-size: 0.8em;
	}

	h5 {
		font-size: 0.8em;
	}

	h6 {
		font-size: 0.8em;
	}

	@media screen and (max-width: 980px) {

		h1 br, h2 br, h3 br, h4 br, h5 br, h6 br {
			display: none;
		}

	}

	@media screen and (max-width: 736px) {

		h2 {
			font-size: 1em;
		}

		h3 {
			font-size: 0.8em;
		}

	}

	sub {
		font-size: 0.8em;
		position: relative;
		top: 0.5em;
	}

	sup {
		font-size: 0.8em;
		position: relative;
		top: -0.5em;
	}

	blockquote {
		border-left: solid 4px #c9c9c9;
		font-style: italic;
		margin: 0 0 2em 0;
		padding: 0.5em 0 0.5em 2em;
	}

	code {
		background: rgba(144, 144, 144, 0.075);
		border-radius: 4px;
		border: solid 1px #c9c9c9;
		font-family: "Courier New", monospace;
		font-size: 0.9em;
		margin: 0 0.25em;
		padding: 0.25em 0.65em;
	}

	pre {
		-webkit-overflow-scrolling: touch;
		font-family: "Courier New", monospace;
		font-size: 0.9em;
		margin: 0 0 2em 0;
	}

		pre code {
			display: block;
			line-height: 1.75;
			padding: 1em 1.5em;
			overflow-x: auto;
		}

	hr {
		border: 0;
		border-bottom: solid 1px #c9c9c9;
		margin: 2em 0;
	}

		hr.major {
			margin: 3em 0;
		}

	.align-left {
		text-align: left;
	}

	.align-center {
		text-align: center;
	}

	.align-right {
		text-align: right;
	}

/* Row */

	.row {
		display: flex;
		flex-wrap: wrap;
		box-sizing: border-box;
		align-items: stretch;
	}

		.row > * {
			box-sizing: border-box;
		}

		.row.gtr-uniform > * > :last-child {
			margin-bottom: 0;
		}

		.row.aln-left {
			justify-content: flex-start;
		}

		.row.aln-center {
			justify-content: center;
		}

		.row.aln-right {
			justify-content: flex-end;
		}

		.row.aln-top {
			align-items: flex-start;
		}

		.row.aln-middle {
			align-items: center;
		}

		.row.aln-bottom {
			align-items: flex-end;
		}

		.row > .imp {
			order: -1;
		}

		.row > .col-1 {
			width: 8.33333%;
		}

		.row > .off-1 {
			margin-left: 8.33333%;
		}

		.row > .col-2 {
			width: 16.66667%;
		}

		.row > .off-2 {
			margin-left: 16.66667%;
		}

		.row > .col-3 {
			width: 25%;
		}

		.row > .off-3 {
			margin-left: 25%;
		}

		.row > .col-4 {
			width: 33.33333%;
		}

		.row > .off-4 {
			margin-left: 33.33333%;
		}

		.row > .col-5 {
			width: 41.66667%;
		}

		.row > .off-5 {
			margin-left: 41.66667%;
		}

		.row > .col-6 {
			width: 50%;
		}

		.row > .off-6 {
			margin-left: 50%;
		}

		.row > .col-7 {
			width: 58.33333%;
		}

		.row > .off-7 {
			margin-left: 58.33333%;
		}

		.row > .col-8 {
			width: 66.66667%;
		}

		.row > .off-8 {
			margin-left: 66.66667%;
		}

		.row > .col-9 {
			width: 75%;
		}

		.row > .off-9 {
			margin-left: 75%;
		}

		.row > .col-10 {
			width: 83.33333%;
		}

		.row > .off-10 {
			margin-left: 83.33333%;
		}

		.row > .col-11 {
			width: 91.66667%;
		}

		.row > .off-11 {
			margin-left: 91.66667%;
		}

		.row > .col-12 {
			width: 100%;
		}

		.row > .off-12 {
			margin-left: 100%;
		}

		.row.gtr-0 {
			margin-top: 0;
			margin-left: 0em;
		}

			.row.gtr-0 > * {
				padding: 0 0 0 0em;
			}

			.row.gtr-0.gtr-uniform {
				margin-top: 0em;
			}

				.row.gtr-0.gtr-uniform > * {
					padding-top: 0em;
				}

		.row.gtr-25 {
			margin-top: 0;
			margin-left: -0.5em;
		}

			.row.gtr-25 > * {
				padding: 0 0 0 0.5em;
			}

			.row.gtr-25.gtr-uniform {
				margin-top: -0.5em;
			}

				.row.gtr-25.gtr-uniform > * {
					padding-top: 0.5em;
				}

		.row.gtr-50 {
			margin-top: 0;
			margin-left: -1em;
		}

			.row.gtr-50 > * {
				padding: 0 0 0 1em;
			}

			.row.gtr-50.gtr-uniform {
				margin-top: -1em;
			}

				.row.gtr-50.gtr-uniform > * {
					padding-top: 1em;
				}

		.row {
			margin-top: 0;
			margin-left: -2em;
		}

			.row > * {
				padding: 0 0 0 2em;
			}

			.row.gtr-uniform {
				margin-top: -2em;
			}

				.row.gtr-uniform > * {
					padding-top: 2em;
				}

		.row.gtr-150 {
			margin-top: 0;
			margin-left: -3em;
		}

			.row.gtr-150 > * {
				padding: 0 0 0 3em;
			}

			.row.gtr-150.gtr-uniform {
				margin-top: -3em;
			}

				.row.gtr-150.gtr-uniform > * {
					padding-top: 3em;
				}

		.row.gtr-200 {
			margin-top: 0;
			margin-left: -4em;
		}

			.row.gtr-200 > * {
				padding: 0 0 0 4em;
			}

			.row.gtr-200.gtr-uniform {
				margin-top: -4em;
			}

				.row.gtr-200.gtr-uniform > * {
					padding-top: 4em;
				}

		@media screen and (max-width: 1680px) {

			.row {
				display: flex;
				flex-wrap: wrap;
				box-sizing: border-box;
				align-items: stretch;
			}

				.row > * {
					box-sizing: border-box;
				}

				.row.gtr-uniform > * > :last-child {
					margin-bottom: 0;
				}

				.row.aln-left {
					justify-content: flex-start;
				}

				.row.aln-center {
					justify-content: center;
				}

				.row.aln-right {
					justify-content: flex-end;
				}

				.row.aln-top {
					align-items: flex-start;
				}

				.row.aln-middle {
					align-items: center;
				}

				.row.aln-bottom {
					align-items: flex-end;
				}

				.row > .imp-xlarge {
					order: -1;
				}

				.row > .col-1-xlarge {
					width: 8.33333%;
				}

				.row > .off-1-xlarge {
					margin-left: 8.33333%;
				}

				.row > .col-2-xlarge {
					width: 16.66667%;
				}

				.row > .off-2-xlarge {
					margin-left: 16.66667%;
				}

				.row > .col-3-xlarge {
					width: 25%;
				}

				.row > .off-3-xlarge {
					margin-left: 25%;
				}

				.row > .col-4-xlarge {
					width: 33.33333%;
				}

				.row > .off-4-xlarge {
					margin-left: 33.33333%;
				}

				.row > .col-5-xlarge {
					width: 41.66667%;
				}

				.row > .off-5-xlarge {
					margin-left: 41.66667%;
				}

				.row > .col-6-xlarge {
					width: 50%;
				}

				.row > .off-6-xlarge {
					margin-left: 50%;
				}

				.row > .col-7-xlarge {
					width: 58.33333%;
				}

				.row > .off-7-xlarge {
					margin-left: 58.33333%;
				}

				.row > .col-8-xlarge {
					width: 66.66667%;
				}

				.row > .off-8-xlarge {
					margin-left: 66.66667%;
				}

				.row > .col-9-xlarge {
					width: 75%;
				}

				.row > .off-9-xlarge {
					margin-left: 75%;
				}

				.row > .col-10-xlarge {
					width: 83.33333%;
				}

				.row > .off-10-xlarge {
					margin-left: 83.33333%;
				}

				.row > .col-11-xlarge {
					width: 91.66667%;
				}

				.row > .off-11-xlarge {
					margin-left: 91.66667%;
				}

				.row > .col-12-xlarge {
					width: 100%;
				}

				.row > .off-12-xlarge {
					margin-left: 100%;
				}

				.row.gtr-0 {
					margin-top: 0;
					margin-left: 0em;
				}

					.row.gtr-0 > * {
						padding: 0 0 0 0em;
					}

					.row.gtr-0.gtr-uniform {
						margin-top: 0em;
					}

						.row.gtr-0.gtr-uniform > * {
							padding-top: 0em;
						}

				.row.gtr-25 {
					margin-top: 0;
					margin-left: -0.5em;
				}

					.row.gtr-25 > * {
						padding: 0 0 0 0.5em;
					}

					.row.gtr-25.gtr-uniform {
						margin-top: -0.5em;
					}

						.row.gtr-25.gtr-uniform > * {
							padding-top: 0.5em;
						}

				.row.gtr-50 {
					margin-top: 0;
					margin-left: -1em;
				}

					.row.gtr-50 > * {
						padding: 0 0 0 1em;
					}

					.row.gtr-50.gtr-uniform {
						margin-top: -1em;
					}

						.row.gtr-50.gtr-uniform > * {
							padding-top: 1em;
						}

				.row {
					margin-top: 0;
					margin-left: -2em;
				}

					.row > * {
						padding: 0 0 0 2em;
					}

					.row.gtr-uniform {
						margin-top: -2em;
					}

						.row.gtr-uniform > * {
							padding-top: 2em;
						}

				.row.gtr-150 {
					margin-top: 0;
					margin-left: -3em;
				}

					.row.gtr-150 > * {
						padding: 0 0 0 3em;
					}

					.row.gtr-150.gtr-uniform {
						margin-top: -3em;
					}

						.row.gtr-150.gtr-uniform > * {
							padding-top: 3em;
						}

				.row.gtr-200 {
					margin-top: 0;
					margin-left: -4em;
				}

					.row.gtr-200 > * {
						padding: 0 0 0 4em;
					}

					.row.gtr-200.gtr-uniform {
						margin-top: -4em;
					}

						.row.gtr-200.gtr-uniform > * {
							padding-top: 4em;
						}

		}

		@media screen and (max-width: 1280px) {

			.row {
				display: flex;
				flex-wrap: wrap;
				box-sizing: border-box;
				align-items: stretch;
			}

				.row > * {
					box-sizing: border-box;
				}

				.row.gtr-uniform > * > :last-child {
					margin-bottom: 0;
				}

				.row.aln-left {
					justify-content: flex-start;
				}

				.row.aln-center {
					justify-content: center;
				}

				.row.aln-right {
					justify-content: flex-end;
				}

				.row.aln-top {
					align-items: flex-start;
				}

				.row.aln-middle {
					align-items: center;
				}

				.row.aln-bottom {
					align-items: flex-end;
				}

				.row > .imp-large {
					order: -1;
				}

				.row > .col-1-large {
					width: 8.33333%;
				}

				.row > .off-1-large {
					margin-left: 8.33333%;
				}

				.row > .col-2-large {
					width: 16.66667%;
				}

				.row > .off-2-large {
					margin-left: 16.66667%;
				}

				.row > .col-3-large {
					width: 25%;
				}

				.row > .off-3-large {
					margin-left: 25%;
				}

				.row > .col-4-large {
					width: 33.33333%;
				}

				.row > .off-4-large {
					margin-left: 33.33333%;
				}

				.row > .col-5-large {
					width: 41.66667%;
				}

				.row > .off-5-large {
					margin-left: 41.66667%;
				}

				.row > .col-6-large {
					width: 50%;
				}

				.row > .off-6-large {
					margin-left: 50%;
				}

				.row > .col-7-large {
					width: 58.33333%;
				}

				.row > .off-7-large {
					margin-left: 58.33333%;
				}

				.row > .col-8-large {
					width: 66.66667%;
				}

				.row > .off-8-large {
					margin-left: 66.66667%;
				}

				.row > .col-9-large {
					width: 75%;
				}

				.row > .off-9-large {
					margin-left: 75%;
				}

				.row > .col-10-large {
					width: 83.33333%;
				}

				.row > .off-10-large {
					margin-left: 83.33333%;
				}

				.row > .col-11-large {
					width: 91.66667%;
				}

				.row > .off-11-large {
					margin-left: 91.66667%;
				}

				.row > .col-12-large {
					width: 100%;
				}

				.row > .off-12-large {
					margin-left: 100%;
				}

				.row.gtr-0 {
					margin-top: 0;
					margin-left: 0em;
				}

					.row.gtr-0 > * {
						padding: 0 0 0 0em;
					}

					.row.gtr-0.gtr-uniform {
						margin-top: 0em;
					}

						.row.gtr-0.gtr-uniform > * {
							padding-top: 0em;
						}

				.row.gtr-25 {
					margin-top: 0;
					margin-left: -0.5em;
				}

					.row.gtr-25 > * {
						padding: 0 0 0 0.5em;
					}

					.row.gtr-25.gtr-uniform {
						margin-top: -0.5em;
					}

						.row.gtr-25.gtr-uniform > * {
							padding-top: 0.5em;
						}

				.row.gtr-50 {
					margin-top: 0;
					margin-left: -1em;
				}

					.row.gtr-50 > * {
						padding: 0 0 0 1em;
					}

					.row.gtr-50.gtr-uniform {
						margin-top: -1em;
					}

						.row.gtr-50.gtr-uniform > * {
							padding-top: 1em;
						}

				.row {
					margin-top: 0;
					margin-left: -2em;
				}

					.row > * {
						padding: 0 0 0 2em;
					}

					.row.gtr-uniform {
						margin-top: -2em;
					}

						.row.gtr-uniform > * {
							padding-top: 2em;
						}

				.row.gtr-150 {
					margin-top: 0;
					margin-left: -3em;
				}

					.row.gtr-150 > * {
						padding: 0 0 0 3em;
					}

					.row.gtr-150.gtr-uniform {
						margin-top: -3em;
					}

						.row.gtr-150.gtr-uniform > * {
							padding-top: 3em;
						}

				.row.gtr-200 {
					margin-top: 0;
					margin-left: -4em;
				}

					.row.gtr-200 > * {
						padding: 0 0 0 4em;
					}

					.row.gtr-200.gtr-uniform {
						margin-top: -4em;
					}

						.row.gtr-200.gtr-uniform > * {
							padding-top: 4em;
						}

		}

		@media screen and (max-width: 980px) {

			.row {
				display: flex;
				flex-wrap: wrap;
				box-sizing: border-box;
				align-items: stretch;
			}

				.row > * {
					box-sizing: border-box;
				}

				.row.gtr-uniform > * > :last-child {
					margin-bottom: 0;
				}

				.row.aln-left {
					justify-content: flex-start;
				}

				.row.aln-center {
					justify-content: center;
				}

				.row.aln-right {
					justify-content: flex-end;
				}

				.row.aln-top {
					align-items: flex-start;
				}

				.row.aln-middle {
					align-items: center;
				}

				.row.aln-bottom {
					align-items: flex-end;
				}

				.row > .imp-medium {
					order: -1;
				}

				.row > .col-1-medium {
					width: 8.33333%;
				}

				.row > .off-1-medium {
					margin-left: 8.33333%;
				}

				.row > .col-2-medium {
					width: 16.66667%;
				}

				.row > .off-2-medium {
					margin-left: 16.66667%;
				}

				.row > .col-3-medium {
					width: 25%;
				}

				.row > .off-3-medium {
					margin-left: 25%;
				}

				.row > .col-4-medium {
					width: 33.33333%;
				}

				.row > .off-4-medium {
					margin-left: 33.33333%;
				}

				.row > .col-5-medium {
					width: 41.66667%;
				}

				.row > .off-5-medium {
					margin-left: 41.66667%;
				}

				.row > .col-6-medium {
					width: 50%;
				}

				.row > .off-6-medium {
					margin-left: 50%;
				}

				.row > .col-7-medium {
					width: 58.33333%;
				}

				.row > .off-7-medium {
					margin-left: 58.33333%;
				}

				.row > .col-8-medium {
					width: 66.66667%;
				}

				.row > .off-8-medium {
					margin-left: 66.66667%;
				}

				.row > .col-9-medium {
					width: 75%;
				}

				.row > .off-9-medium {
					margin-left: 75%;
				}

				.row > .col-10-medium {
					width: 83.33333%;
				}

				.row > .off-10-medium {
					margin-left: 83.33333%;
				}

				.row > .col-11-medium {
					width: 91.66667%;
				}

				.row > .off-11-medium {
					margin-left: 91.66667%;
				}

				.row > .col-12-medium {
					width: 100%;
				}

				.row > .off-12-medium {
					margin-left: 100%;
				}

				.row.gtr-0 {
					margin-top: 0;
					margin-left: 0em;
				}

					.row.gtr-0 > * {
						padding: 0 0 0 0em;
					}

					.row.gtr-0.gtr-uniform {
						margin-top: 0em;
					}

						.row.gtr-0.gtr-uniform > * {
							padding-top: 0em;
						}

				.row.gtr-25 {
					margin-top: 0;
					margin-left: -0.375em;
				}

					.row.gtr-25 > * {
						padding: 0 0 0 0.375em;
					}

					.row.gtr-25.gtr-uniform {
						margin-top: -0.375em;
					}

						.row.gtr-25.gtr-uniform > * {
							padding-top: 0.375em;
						}

				.row.gtr-50 {
					margin-top: 0;
					margin-left: -0.75em;
				}

					.row.gtr-50 > * {
						padding: 0 0 0 0.75em;
					}

					.row.gtr-50.gtr-uniform {
						margin-top: -0.75em;
					}

						.row.gtr-50.gtr-uniform > * {
							padding-top: 0.75em;
						}

				.row {
					margin-top: 0;
					margin-left: -1.5em;
				}

					.row > * {
						padding: 0 0 0 1.5em;
					}

					.row.gtr-uniform {
						margin-top: -1.5em;
					}

						.row.gtr-uniform > * {
							padding-top: 1.5em;
						}

				.row.gtr-150 {
					margin-top: 0;
					margin-left: -2.25em;
				}

					.row.gtr-150 > * {
						padding: 0 0 0 2.25em;
					}

					.row.gtr-150.gtr-uniform {
						margin-top: -2.25em;
					}

						.row.gtr-150.gtr-uniform > * {
							padding-top: 2.25em;
						}

				.row.gtr-200 {
					margin-top: 0;
					margin-left: -3em;
				}

					.row.gtr-200 > * {
						padding: 0 0 0 3em;
					}

					.row.gtr-200.gtr-uniform {
						margin-top: -3em;
					}

						.row.gtr-200.gtr-uniform > * {
							padding-top: 3em;
						}

		}

		@media screen and (max-width: 736px) {

			.row {
				display: flex;
				flex-wrap: wrap;
				box-sizing: border-box;
				align-items: stretch;
			}

				.row > * {
					box-sizing: border-box;
				}

				.row.gtr-uniform > * > :last-child {
					margin-bottom: 0;
				}

				.row.aln-left {
					justify-content: flex-start;
				}

				.row.aln-center {
					justify-content: center;
				}

				.row.aln-right {
					justify-content: flex-end;
				}

				.row.aln-top {
					align-items: flex-start;
				}

				.row.aln-middle {
					align-items: center;
				}

				.row.aln-bottom {
					align-items: flex-end;
				}

				.row > .imp-small {
					order: -1;
				}

				.row > .col-1-small {
					width: 8.33333%;
				}

				.row > .off-1-small {
					margin-left: 8.33333%;
				}

				.row > .col-2-small {
					width: 16.66667%;
				}

				.row > .off-2-small {
					margin-left: 16.66667%;
				}

				.row > .col-3-small {
					width: 25%;
				}

				.row > .off-3-small {
					margin-left: 25%;
				}

				.row > .col-4-small {
					width: 33.33333%;
				}

				.row > .off-4-small {
					margin-left: 33.33333%;
				}

				.row > .col-5-small {
					width: 41.66667%;
				}

				.row > .off-5-small {
					margin-left: 41.66667%;
				}

				.row > .col-6-small {
					width: 50%;
				}

				.row > .off-6-small {
					margin-left: 50%;
				}

				.row > .col-7-small {
					width: 58.33333%;
				}

				.row > .off-7-small {
					margin-left: 58.33333%;
				}

				.row > .col-8-small {
					width: 66.66667%;
				}

				.row > .off-8-small {
					margin-left: 66.66667%;
				}

				.row > .col-9-small {
					width: 75%;
				}

				.row > .off-9-small {
					margin-left: 75%;
				}

				.row > .col-10-small {
					width: 83.33333%;
				}

				.row > .off-10-small {
					margin-left: 83.33333%;
				}

				.row > .col-11-small {
					width: 91.66667%;
				}

				.row > .off-11-small {
					margin-left: 91.66667%;
				}

				.row > .col-12-small {
					width: 100%;
				}

				.row > .off-12-small {
					margin-left: 100%;
				}

				.row.gtr-0 {
					margin-top: 0;
					margin-left: 0em;
				}

					.row.gtr-0 > * {
						padding: 0 0 0 0em;
					}

					.row.gtr-0.gtr-uniform {
						margin-top: 0em;
					}

						.row.gtr-0.gtr-uniform > * {
							padding-top: 0em;
						}

				.row.gtr-25 {
					margin-top: 0;
					margin-left: -0.25em;
				}

					.row.gtr-25 > * {
						padding: 0 0 0 0.25em;
					}

					.row.gtr-25.gtr-uniform {
						margin-top: -0.25em;
					}

						.row.gtr-25.gtr-uniform > * {
							padding-top: 0.25em;
						}

				.row.gtr-50 {
					margin-top: 0;
					margin-left: -0.5em;
				}

					.row.gtr-50 > * {
						padding: 0 0 0 0.5em;
					}

					.row.gtr-50.gtr-uniform {
						margin-top: -0.5em;
					}

						.row.gtr-50.gtr-uniform > * {
							padding-top: 0.5em;
						}

				.row {
					margin-top: 0;
					margin-left: -1em;
				}

					.row > * {
						padding: 0 0 0 1em;
					}

					.row.gtr-uniform {
						margin-top: -1em;
					}

						.row.gtr-uniform > * {
							padding-top: 1em;
						}

				.row.gtr-150 {
					margin-top: 0;
					margin-left: -1.5em;
				}

					.row.gtr-150 > * {
						padding: 0 0 0 1.5em;
					}

					.row.gtr-150.gtr-uniform {
						margin-top: -1.5em;
					}

						.row.gtr-150.gtr-uniform > * {
							padding-top: 1.5em;
						}

				.row.gtr-200 {
					margin-top: 0;
					margin-left: -2em;
				}

					.row.gtr-200 > * {
						padding: 0 0 0 2em;
					}

					.row.gtr-200.gtr-uniform {
						margin-top: -2em;
					}

						.row.gtr-200.gtr-uniform > * {
							padding-top: 2em;
						}

		}

		@media screen and (max-width: 480px) {

			.row {
				display: flex;
				flex-wrap: wrap;
				box-sizing: border-box;
				align-items: stretch;
			}

				.row > * {
					box-sizing: border-box;
				}

				.row.gtr-uniform > * > :last-child {
					margin-bottom: 0;
				}

				.row.aln-left {
					justify-content: flex-start;
				}

				.row.aln-center {
					justify-content: center;
				}

				.row.aln-right {
					justify-content: flex-end;
				}

				.row.aln-top {
					align-items: flex-start;
				}

				.row.aln-middle {
					align-items: center;
				}

				.row.aln-bottom {
					align-items: flex-end;
				}

				.row > .imp-xsmall {
					order: -1;
				}

				.row > .col-1-xsmall {
					width: 8.33333%;
				}

				.row > .off-1-xsmall {
					margin-left: 8.33333%;
				}

				.row > .col-2-xsmall {
					width: 16.66667%;
				}

				.row > .off-2-xsmall {
					margin-left: 16.66667%;
				}

				.row > .col-3-xsmall {
					width: 25%;
				}

				.row > .off-3-xsmall {
					margin-left: 25%;
				}

				.row > .col-4-xsmall {
					width: 33.33333%;
				}

				.row > .off-4-xsmall {
					margin-left: 33.33333%;
				}

				.row > .col-5-xsmall {
					width: 41.66667%;
				}

				.row > .off-5-xsmall {
					margin-left: 41.66667%;
				}

				.row > .col-6-xsmall {
					width: 50%;
				}

				.row > .off-6-xsmall {
					margin-left: 50%;
				}

				.row > .col-7-xsmall {
					width: 58.33333%;
				}

				.row > .off-7-xsmall {
					margin-left: 58.33333%;
				}

				.row > .col-8-xsmall {
					width: 66.66667%;
				}

				.row > .off-8-xsmall {
					margin-left: 66.66667%;
				}

				.row > .col-9-xsmall {
					width: 75%;
				}

				.row > .off-9-xsmall {
					margin-left: 75%;
				}

				.row > .col-10-xsmall {
					width: 83.33333%;
				}

				.row > .off-10-xsmall {
					margin-left: 83.33333%;
				}

				.row > .col-11-xsmall {
					width: 91.66667%;
				}

				.row > .off-11-xsmall {
					margin-left: 91.66667%;
				}

				.row > .col-12-xsmall {
					width: 100%;
				}

				.row > .off-12-xsmall {
					margin-left: 100%;
				}

				.row.gtr-0 {
					margin-top: 0;
					margin-left: 0em;
				}

					.row.gtr-0 > * {
						padding: 0 0 0 0em;
					}

					.row.gtr-0.gtr-uniform {
						margin-top: 0em;
					}

						.row.gtr-0.gtr-uniform > * {
							padding-top: 0em;
						}

				.row.gtr-25 {
					margin-top: 0;
					margin-left: -0.25em;
				}

					.row.gtr-25 > * {
						padding: 0 0 0 0.25em;
					}

					.row.gtr-25.gtr-uniform {
						margin-top: -0.25em;
					}

						.row.gtr-25.gtr-uniform > * {
							padding-top: 0.25em;
						}

				.row.gtr-50 {
					margin-top: 0;
					margin-left: -0.5em;
				}

					.row.gtr-50 > * {
						padding: 0 0 0 0.5em;
					}

					.row.gtr-50.gtr-uniform {
						margin-top: -0.5em;
					}

						.row.gtr-50.gtr-uniform > * {
							padding-top: 0.5em;
						}

				.row {
					margin-top: 0;
					margin-left: -1em;
				}

					.row > * {
						padding: 0 0 0 1em;
					}

					.row.gtr-uniform {
						margin-top: -1em;
					}

						.row.gtr-uniform > * {
							padding-top: 1em;
						}

				.row.gtr-150 {
					margin-top: 0;
					margin-left: -1.5em;
				}

					.row.gtr-150 > * {
						padding: 0 0 0 1.5em;
					}

					.row.gtr-150.gtr-uniform {
						margin-top: -1.5em;
					}

						.row.gtr-150.gtr-uniform > * {
							padding-top: 1.5em;
						}

				.row.gtr-200 {
					margin-top: 0;
					margin-left: -2em;
				}

					.row.gtr-200 > * {
						padding: 0 0 0 2em;
					}

					.row.gtr-200.gtr-uniform {
						margin-top: -2em;
					}

						.row.gtr-200.gtr-uniform > * {
							padding-top: 2em;
						}

		}

/* Section/Article */

	section.special, article.special {
		text-align: center;
	}

	header p {
		margin-top: -1em;
	}

	@media screen and (max-width: 736px) {

		header p {
			margin-top: 0;
		}

	}

/* Icon */

	.icon {
		text-decoration: none;
		border-bottom: none;
		position: relative;
	}

		.icon:before {
			-moz-osx-font-smoothing: grayscale;
			-webkit-font-smoothing: antialiased;
			display: inline-block;
			font-style: normal;
			font-variant: normal;
			text-rendering: auto;
			line-height: 1;
			text-transform: none !important;
			font-family: 'Font Awesome 5 Free';
			font-weight: 400;
		}

		.icon > .label {
			display: none;
		}

		.icon:before {
			line-height: inherit;
		}

		.icon.solid:before {
			font-weight: 900;
		}

		.icon.brands:before {
			font-family: 'Font Awesome 5 Brands';
		}

		.icon.style2 {
			-moz-transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out, border-color 0.2s ease-in-out;
			-webkit-transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out, border-color 0.2s ease-in-out;
			-ms-transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out, border-color 0.2s ease-in-out;
			transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out, border-color 0.2s ease-in-out;
			background-color: transparent;
			border: solid 1px #c9c9c9;
			border-radius: 4px;
			width: 2.65em;
			height: 2.65em;
			display: inline-block;
			text-align: center;
			line-height: 2.65em;
			color: inherit;
		}

			.icon.style2:before {
				font-size: 1.1em;
			}

			.icon.style2:hover {
				color: #f2849e;
				border-color: #f2849e;
			}

			.icon.style2:active {
				background-color: rgba(242, 132, 158, 0.1);
			}

/* List */

	ol {
		list-style: decimal;
		margin: 0 0 2em 0;
		padding-left: 1.25em;
	}

		ol li {
			padding-left: 0.25em;
		}

	ul {
		list-style: disc;
		margin: 0 0 2em 0;
		padding-left: 1em;
	}

		ul li {
			padding-left: 0.5em;
		}

		ul.alt {
			list-style: none;
			padding-left: 0;
		}

			ul.alt li {
				border-top: solid 1px #c9c9c9;
				padding: 0.5em 0;
			}

				ul.alt li:first-child {
					border-top: 0;
					padding-top: 0;
				}

	dl {
		margin: 0 0 2em 0;
	}

		dl dt {
			display: block;
			font-weight: 900;
			margin: 0 0 1em 0;
		}

		dl dd {
			margin-left: 2em;
		}

/* Actions */

	ul.actions {
		display: -moz-flex;
		display: -webkit-flex;
		display: -ms-flex;
		display: flex;
		cursor: default;
		list-style: none;
		margin-left: -1em;
		padding-left: 0;
	}

		ul.actions li {
			padding: 0 0 0 1em;
			vertical-align: middle;
		}

		ul.actions.special {
			-moz-justify-content: center;
			-webkit-justify-content: center;
			-ms-justify-content: center;
			justify-content: center;
			width: 100%;
			margin-left: 0;
		}

			ul.actions.special li:first-child {
				padding-left: 0;
			}

		ul.actions.stacked {
			-moz-flex-direction: column;
			-webkit-flex-direction: column;
			-ms-flex-direction: column;
			flex-direction: column;
			margin-left: 0;
		}

			ul.actions.stacked li {
				padding: 1.3em 0 0 0;
			}

				ul.actions.stacked li:first-child {
					padding-top: 0;
				}

		ul.actions.fit {
			width: calc(100% + 1em);
		}

			ul.actions.fit li {
				-moz-flex-grow: 1;
				-webkit-flex-grow: 1;
				-ms-flex-grow: 1;
				flex-grow: 1;
				-moz-flex-shrink: 1;
				-webkit-flex-shrink: 1;
				-ms-flex-shrink: 1;
				flex-shrink: 1;
				width: 100%;
			}

				ul.actions.fit li > * {
					width: 100%;
				}

			ul.actions.fit.stacked {
				width: 100%;
			}

		@media screen and (max-width: 480px) {

			ul.actions:not(.fixed) {
				-moz-flex-direction: column;
				-webkit-flex-direction: column;
				-ms-flex-direction: column;
				flex-direction: column;
				margin-left: 0;
				width: 100% !important;
			}

				ul.actions:not(.fixed) li {
					-moz-flex-grow: 1;
					-webkit-flex-grow: 1;
					-ms-flex-grow: 1;
					flex-grow: 1;
					-moz-flex-shrink: 1;
					-webkit-flex-shrink: 1;
					-ms-flex-shrink: 1;
					flex-shrink: 1;
					padding: 1em 0 0 0;
					text-align: center;
					width: 100%;
				}

					ul.actions:not(.fixed) li > * {
						width: 100%;
					}

					ul.actions:not(.fixed) li:first-child {
						padding-top: 0;
					}

					ul.actions:not(.fixed) li input[type="submit"],
					ul.actions:not(.fixed) li input[type="reset"],
					ul.actions:not(.fixed) li input[type="button"],
					ul.actions:not(.fixed) li button,
					ul.actions:not(.fixed) li .button {
						width: 100%;
					}

						ul.actions:not(.fixed) li input[type="submit"].icon:before,
						ul.actions:not(.fixed) li input[type="reset"].icon:before,
						ul.actions:not(.fixed) li input[type="button"].icon:before,
						ul.actions:not(.fixed) li button.icon:before,
						ul.actions:not(.fixed) li .button.icon:before {
							margin-left: -0.5rem;
						}

		}

/* Icons */

	ul.icons {
		cursor: default;
		list-style: none;
		padding-left: 0;
		margin: -1em 0 2em -1em;
	}

		ul.icons li {
			display: inline-block;
			padding: 1em 0 0 1em;
		}

/* Form */

	form {
		margin: 0 0 2em 0;
		overflow-x: hidden;
	}

		form > :last-child {
			margin-bottom: 0;
		}

		form > .fields {
			display: -moz-flex;
			display: -webkit-flex;
			display: -ms-flex;
			display: flex;
			-moz-flex-wrap: wrap;
			-webkit-flex-wrap: wrap;
			-ms-flex-wrap: wrap;
			flex-wrap: wrap;
			width: calc(100% + 3em);
			margin: -1.5em 0 2em -1.5em;
		}

			form > .fields > .field {
				-moz-flex-grow: 0;
				-webkit-flex-grow: 0;
				-ms-flex-grow: 0;
				flex-grow: 0;
				-moz-flex-shrink: 0;
				-webkit-flex-shrink: 0;
				-ms-flex-shrink: 0;
				flex-shrink: 0;
				padding: 1.5em 0 0 1.5em;
				width: calc(100% - 1.5em);
			}

				form > .fields > .field.half {
					width: calc(50% - 0.75em);
				}

				form > .fields > .field.third {
					width: calc(100%/3 - 0.5em);
				}

				form > .fields > .field.quarter {
					width: calc(25% - 0.375em);
				}

		@media screen and (max-width: 480px) {

			form > .fields {
				width: calc(100% + 3em);
				margin: -1.5em 0 2em -1.5em;
			}

				form > .fields > .field {
					padding: 1.5em 0 0 1.5em;
					width: calc(100% - 1.5em);
				}

					form > .fields > .field.half {
						width: calc(100% - 1.5em);
					}

					form > .fields > .field.third {
						width: calc(100% - 1.5em);
					}

					form > .fields > .field.quarter {
						width: calc(100% - 1.5em);
					}

		}

	label {
		display: block;
		font-size: 0.9em;
		font-weight: 900;
		margin: 0 0 1em 0;
	}

	input[type="text"],
	input[type="password"],
	input[type="email"],
	input[type="tel"],
	select,
	textarea {
		-moz-appearance: none;
		-webkit-appearance: none;
		-ms-appearance: none;
		appearance: none;
		background-color: transparent;
		border: none;
		border-radius: 0;
		border-bottom: solid 1px #c9c9c9;
		color: inherit;
		display: block;
		outline: 0;
		padding: 0;
		text-decoration: none;
		width: 100%;
	}

		input[type="text"]:invalid,
		input[type="password"]:invalid,
		input[type="email"]:invalid,
		input[type="tel"]:invalid,
		select:invalid,
		textarea:invalid {
			box-shadow: none;
		}

		input[type="text"]:focus,
		input[type="password"]:focus,
		input[type="email"]:focus,
		input[type="tel"]:focus,
		select:focus,
		textarea:focus {
			border-bottom-color: #f2849e;
			box-shadow: inset 0 -1px 0 0 #f2849e;
		}

	select {
		background-size: 1.25rem;
		background-repeat: no-repeat;
		background-position: calc(100% - 1rem) center;
		height: 3em;
		padding-right: 3em;
		text-overflow: ellipsis;
	}

		select option {
			background: #ffffff;
		}

		select:focus::-ms-value {
			background-color: transparent;
		}

		select::-ms-expand {
			display: none;
		}

	input[type="text"],
	input[type="password"],
	input[type="email"],
	select {
		height: 3em;
	}

	textarea {
		padding: 0;
		min-height: 3.75em;
	}

	input[type="checkbox"],
	input[type="radio"] {
		-moz-appearance: none;
		-webkit-appearance: none;
		-ms-appearance: none;
		appearance: none;
		display: block;
		float: left;
		margin-right: -2em;
		opacity: 0;
		width: 1em;
		z-index: -1;
	}

		input[type="checkbox"] + label,
		input[type="radio"] + label {
			text-decoration: none;
			color: #585858;
			cursor: pointer;
			display: inline-block;
			font-size: 1em;
			font-weight: 300;
			padding-left: 2.55em;
			padding-right: 0.75em;
			position: relative;
		}

			input[type="checkbox"] + label:before,
			input[type="radio"] + label:before {
				-moz-osx-font-smoothing: grayscale;
				-webkit-font-smoothing: antialiased;
				display: inline-block;
				font-style: normal;
				font-variant: normal;
				text-rendering: auto;
				line-height: 1;
				text-transform: none !important;
				font-family: 'Font Awesome 5 Free';
				font-weight: 900;
			}

			input[type="checkbox"] + label:before,
			input[type="radio"] + label:before {
				border-radius: 4px;
				border: solid 1px #c9c9c9;
				content: '';
				display: inline-block;
				font-size: 0.8em;
				height: 2.25em;
				left: 0;
				line-height: 2.25em;
				position: absolute;
				text-align: center;
				top: 0;
				width: 2.25em;
			}

		input[type="checkbox"]:checked + label:before,
		input[type="radio"]:checked + label:before {
			background: #585858;
			border-color: #585858;
			color: #ffffff;
			content: '\f00c';
		}

		input[type="checkbox"]:focus + label:before,
		input[type="radio"]:focus + label:before {
			border-color: #f2849e;
			box-shadow: 0 0 0 1px #f2849e;
		}

	input[type="checkbox"] + label:before {
		border-radius: 4px;
	}

	input[type="radio"] + label:before {
		border-radius: 100%;
	}

/* Box */

	.box {
		border-radius: 4px;
		border: solid 1px #c9c9c9;
		margin-bottom: 2em;
		padding: 1.5em;
	}

		.box > :last-child,
		.box > :last-child > :last-child,
		.box > :last-child > :last-child > :last-child {
			margin-bottom: 0;
		}

		.box.alt {
			border: 0;
			border-radius: 0;
			padding: 0;
		}

/* Image */

	.image {
		border-radius: 4px;
		border: 0;
		display: inline-block;
		position: relative;
	}

		.image img {
			border-radius: 4px;
			display: block;
		}

		.image.left, .image.right {
			max-width: 40%;
		}

			.image.left img, .image.right img {
				width: 100%;
			}

		.image.left {
			float: left;
			padding: 0 1.5em 1em 0;
			top: 0.25em;
		}

		.image.right {
			float: right;
			padding: 0 0 1em 1.5em;
			top: 0.25em;
		}

		.image.fit {
			display: block;
			margin: 0 0 2em 0;
			width: 100%;
		}

			.image.fit img {
				width: 100%;
			}

		.image.main {
			display: block;
			margin: 0 0 3em 0;
			width: 100%;
		}

			.image.main img {
				width: 100%;
			}

			@media screen and (max-width: 736px) {

				.image.main {
					margin: 0 0 2em 0;
				}

			}

/* Table */

	.table-wrapper {
		-webkit-overflow-scrolling: touch;
		overflow-x: auto;
	}

	table {
		margin: 0 0 2em 0;
		width: 100%;
	}

		table tbody tr {
			border: solid 1px #c9c9c9;
			border-left: 0;
			border-right: 0;
		}

			table tbody tr:nth-child(2n + 1) {
				background-color: rgba(144, 144, 144, 0.075);
			}

		table td {
			padding: 0.75em 0.75em;
		}

		table th {
			font-size: 0.9em;
			font-weight: 900;
			padding: 0 0.75em 0.75em 0.75em;
			text-align: left;
		}

		table thead {
			border-bottom: solid 2px #c9c9c9;
		}

		table tfoot {
			border-top: solid 2px #c9c9c9;
		}

		table.alt {
			border-collapse: separate;
		}

			table.alt tbody tr td {
				border: solid 1px #c9c9c9;
				border-left-width: 0;
				border-top-width: 0;
			}

				table.alt tbody tr td:first-child {
					border-left-width: 1px;
				}

			table.alt tbody tr:first-child td {
				border-top-width: 1px;
			}

			table.alt thead {
				border-bottom: 0;
			}

			table.alt tfoot {
				border-top: 0;
			}

/* Button */

	input[type="submit"],
	input[type="reset"],
	input[type="button"],
	button,
	.button {
	}

		input[type="submit"].icon:before,
		input[type="reset"].icon:before,
		input[type="button"].icon:before,
		button.icon:before,
		.button.icon:before {
			margin-right: 0.5em;
		}

		input[type="submit"].fit,
		input[type="reset"].fit,
		input[type="button"].fit,
		button.fit,
		.button.fit {
			width: 100%;
		}

		input[type="submit"]:hover,
		input[type="reset"]:hover,
		input[type="button"]:hover,
		button:hover,
		.button:hover {
			color: #f2849e !important;
			box-shadow: inset 0 0 0 2px #f2849e;
		}

		input[type="submit"]:active,
		input[type="reset"]:active,
		input[type="button"]:active,
		button:active,
		.button:active {
			background-color: rgba(242, 132, 158, 0.1);
		}

		input[type="submit"].small,
		input[type="reset"].small,
		input[type="button"].small,
		button.small,
		.button.small {
			font-size: 0.6em;
		}

		input[type="submit"].large,
		input[type="reset"].large,
		input[type="button"].large,
		button.large,
		.button.large {
			font-size: 1em;
		}

		input[type="submit"].primary,
		input[type="reset"].primary,
		input[type="button"].primary,
		button.primary,
		.button.primary {
			box-shadow: none;
			background-color: #585858;
			color: #ffffff !important;
		}

			input[type="submit"].primary:hover,
			input[type="reset"].primary:hover,
			input[type="button"].primary:hover,
			button.primary:hover,
			.button.primary:hover {
				background-color: #f2849e;
			}

			input[type="submit"].primary:active,
			input[type="reset"].primary:active,
			input[type="button"].primary:active,
			button.primary:active,
			.button.primary:active {
				background-color: #ee5f81;
			}

		input[type="submit"].disabled, input[type="submit"]:disabled,
		input[type="reset"].disabled,
		input[type="reset"]:disabled,
		input[type="button"].disabled,
		input[type="button"]:disabled,
		button.disabled,
		button:disabled,
		.button.disabled,
		.button:disabled {
			pointer-events: none;
			opacity: 0.25;
		}
            
    .image::before,
.image::after {
    display: none !important;
    content: none !important;
}

/* Tiles */

	.tiles {
		display: -moz-flex;
		display: -webkit-flex;
		display: -ms-flex;
		display: flex;
		-moz-flex-wrap: wrap;
		-webkit-flex-wrap: wrap;
		-ms-flex-wrap: wrap;
		flex-wrap: wrap;
		postiion: relative;
		margin: -2.5em 0 0 -2.5em;
	}

		.tiles article {
			-moz-transition: -moz-transform 0.5s ease, opacity 0.5s ease;
			-webkit-transition: -webkit-transform 0.5s ease, opacity 0.5s ease;
			-ms-transition: -ms-transform 0.5s ease, opacity 0.5s ease;
			transition: transform 0.5s ease, opacity 0.5s ease;
			position: relative;
			width: calc(33.33333% - 2.5em);
			margin: 2.5em 0 0 2.5em;
		}

			.tiles article > .image {
				-moz-transition: -moz-transform 0.5s ease;
				-webkit-transition: -webkit-transform 0.5s ease;
				-ms-transition: -ms-transform 0.5s ease;
				transition: transform 0.5s ease;
				position: relative;
				display: block;
				width: 100%;
				border-radius: 4px;
				overflow: hidden;
			}

				.tiles article > .image img {
					display: block;
					width: 100%;
				}

				.tiles article > .image:before {
					pointer-events: none;
					-moz-transition: background-color 0.5s ease, opacity 0.5s ease;
					-webkit-transition: background-color 0.5s ease, opacity 0.5s ease;
					-ms-transition: background-color 0.5s ease, opacity 0.5s ease;
					transition: background-color 0.5s ease, opacity 0.5s ease;
					content: '';
					display: block;
					position: absolute;
					top: 0;
					left: 0;
					width: 100%;
					height: 100%;
					opacity: 1.0;
					z-index: 1;
					opacity: 0.8;
				}

				.tiles article > .image:after {
					pointer-events: none;
					-moz-transition: opacity 0.5s ease;
					-webkit-transition: opacity 0.5s ease;
					-ms-transition: opacity 0.5s ease;
					transition: opacity 0.5s ease;
					content: '';
					display: block;
					position: absolute;
					top: 0;
					left: 0;
					width: 100%;
					height: 100%;
					background-position: center;
					background-repeat: no-repeat;
					background-size: 100% 100%;
					opacity: 0.25;
					z-index: 2;
				}

			.tiles article > a {
				display: -moz-flex;
				display: -webkit-flex;
				display: -ms-flex;
				display: flex;
				-moz-flex-direction: column;
				-webkit-flex-direction: column;
				-ms-flex-direction: column;
				flex-direction: column;
				-moz-align-items: center;
				-webkit-align-items: center;
				-ms-align-items: center;
				align-items: center;
				-moz-justify-content: center;
				-webkit-justify-content: center;
				-ms-justify-content: center;
				justify-content: center;
				-moz-transition: background-color 0.5s ease, -moz-transform 0.5s ease;
				-webkit-transition: background-color 0.5s ease, -webkit-transform 0.5s ease;
				-ms-transition: background-color 0.5s ease, -ms-transform 0.5s ease;
				transition: background-color 0.5s ease, transform 0.5s ease;
				position: absolute;
				top: 0;
				left: 0;
				width: 100%;
				height: 100%;
				padding: 1em;
				border-radius: 4px;
				border-bottom: 0;
				color: #ffffff;
				text-align: center;
				text-decoration: none;
				z-index: 3;
			}

				.tiles article > a > :last-child {
					margin: 0;
				}

				.tiles article > a:hover {
					color: #ffffff !important;
				}

				.tiles article > a h2 {
					margin: 0;
				}

				.tiles article > a .content {
					-moz-transition: max-height 0.5s ease, opacity 0.5s ease;
					-webkit-transition: max-height 0.5s ease, opacity 0.5s ease;
					-ms-transition: max-height 0.5s ease, opacity 0.5s ease;
					transition: max-height 0.5s ease, opacity 0.5s ease;
					width: 100%;
					max-height: 0;
					line-height: 1.5;
					margin-top: 0.35em;
					opacity: 0;
				}

					.tiles article > a .content > :last-child {
						margin-bottom: 0;
					}

			.tiles article.style1 > .image:before {
				background-color: #f2849e;
			}

			.tiles article.style2 > .image:before {
				background-color: #7ecaf6;
			}

			.tiles article.style3 > .image:before {
				background-color: #7bd0c1;
			}

			.tiles article.style4 > .image:before {
				background-color: #c75b9b;
			}

			.tiles article.style5 > .image:before {
				background-color: #ae85ca;
			}

			.tiles article.style6 > .image:before {
				background-color: #8499e7;
			}

			body:not(.is-touch) .tiles article:hover > .image {
				-moz-transform: scale(1.1);
				-webkit-transform: scale(1.1);
				-ms-transform: scale(1.1);
				transform: scale(1.1);
			}

				body:not(.is-touch) .tiles article:hover > .image:before {
					background-color: #333333;
					opacity: 0.35;
				}

				body:not(.is-touch) .tiles article:hover > .image:after {
					opacity: 0;
				}

			body:not(.is-touch) .tiles article:hover .content {
				max-height: 15em;
				opacity: 1;
			}

		* + .tiles {
			margin-top: 2em;
		}

		body.is-preload .tiles article {
			-moz-transform: scale(0.9);
			-webkit-transform: scale(0.9);
			-ms-transform: scale(0.9);
			transform: scale(0.9);
			opacity: 0;
		}

		body.is-touch .tiles article .content {
			max-height: 15em;
			opacity: 1;
		}

		@media screen and (max-width: 1280px) {

			.tiles {
				margin: -1.25em 0 0 -1.25em;
			}

				.tiles article {
					width: calc(33.33333% - 1.25em);
					margin: 1.25em 0 0 1.25em;
				}

		}

		@media screen and (max-width: 980px) {

			.tiles {
				margin: -2.5em 0 0 -2.5em;
			}

				.tiles article {
					width: calc(50% - 2.5em);
					margin: 2.5em 0 0 2.5em;
				}

		}

		@media screen and (max-width: 736px) {

			.tiles {
				margin: -1.25em 0 0 -1.25em;
			}

				.tiles article {
					width: calc(50% - 1.25em);
					margin: 1.25em 0 0 1.25em;
				}

					.tiles article:hover > .image {
						-moz-transform: scale(1.0);
						-webkit-transform: scale(1.0);
						-ms-transform: scale(1.0);
						transform: scale(1.0);
					}

		}

		@media screen and (max-width: 480px) {

			.tiles {
				margin: 0;
			}

				.tiles article {
					width: 100%;
					margin: 1.25em 0 0 0;
				}

		}

/* Header */

		#header .logo {
			display: block;
			border-bottom: 0;
			color: inherit;
			font-weight: 900;
			letter-spacing: 0.35em;
			margin: 0 0 2.5em 0;
			text-decoration: none;
			text-transform: uppercase;
			display: inline-block;
		}

			#header .logo > * {
				display: inline-block;
				vertical-align: middle;
			}

			#header .logo .symbol {
				margin-right: 0.65em;
			}

				#header .logo .symbol img {
					display: block;
					width: 2em;
					height: 2em;
				}

		#header nav {
			position: fixed;
			right: 2em;
			top: 2em;
			z-index: 10000;
		}

			#header nav ul {
				display: -moz-flex;
				display: -webkit-flex;
				display: -ms-flex;
				display: flex;
				-moz-align-items: center;
				-webkit-align-items: center;
				-ms-align-items: center;
				align-items: center;
				list-style: none;
				margin: 0;
				padding: 0;
			}

				#header nav ul li {
					display: block;
					padding: 0;
				}

					#header nav ul li a {
						display: block;
						position: relative;
						height: 3em;
						line-height: 3em;
						padding: 0 1.5em;
						background-color: rgba(255, 255, 255, 0.5);
						border-radius: 4px;
						border: 0;
						font-size: 0.8em;
						font-weight: 900;
						letter-spacing: 0.35em;
						text-transform: uppercase;
					}

					#header nav ul li a[href="#menu"] {
						-webkit-tap-highlight-color: transparent;
						width: 4em;
						text-indent: 4em;
						font-size: 1em;
						overflow: hidden;
						padding: 0;
						white-space: nowrap;
					}

						#header nav ul li a[href="#menu"]:before, #header nav ul li a[href="#menu"]:after {
							-moz-transition: opacity 0.2s ease;
							-webkit-transition: opacity 0.2s ease;
							-ms-transition: opacity 0.2s ease;
							transition: opacity 0.2s ease;
							content: '';
							display: block;
							position: absolute;
							top: 0;
							left: 0;
							width: 100%;
							height: 100%;
							background-position: center;
							background-repeat: no-repeat;
							background-size: 2em 2em;
						}

						#header nav ul li a[href="#menu"]:before {
							opacity: 0;
						}

						#header nav ul li a[href="#menu"]:after {
							opacity: 1;
						}

						#header nav ul li a[href="#menu"]:hover:before {
							opacity: 1;
						}

						#header nav ul li a[href="#menu"]:hover:after {
							opacity: 0;
						}

		@media screen and (max-width: 736px) {

			#header {
				padding: 4em 0 0.1em 0 ;
			}

				#header nav {
					right: 0.5em;
					top: 0.5em;
				}

					#header nav ul li a[href="#menu"]:before, #header nav ul li a[href="#menu"]:after {
						background-size: 1.5em 1.5em;
					}

		}

/* Menu */

	#wrapper {
		-moz-transition: opacity 0.45s ease;
		-webkit-transition: opacity 0.45s ease;
		-ms-transition: opacity 0.45s ease;
		transition: opacity 0.45s ease;
		opacity: 1;
	}

	#menu {
		-moz-transform: translateX(22em);
		-webkit-transform: translateX(22em);
		-ms-transform: translateX(22em);
		transform: translateX(22em);
		-moz-transition: -moz-transform 0.45s ease, visibility 0.45s;
		-webkit-transition: -webkit-transform 0.45s ease, visibility 0.45s;
		-ms-transition: -ms-transform 0.45s ease, visibility 0.45s;
		transition: transform 0.45s ease, visibility 0.45s;
		position: fixed;
		top: 0;
		right: 0;
		width: 22em;
		max-width: 80%;
		height: 100%;
		-webkit-overflow-scrolling: touch;
		background: #585858;
		color: #ffffff;
		cursor: default;
		visibility: hidden;
		z-index: 10002;
	}

		#menu > .inner {
			-moz-transition: opacity 0.45s ease;
			-webkit-transition: opacity 0.45s ease;
			-ms-transition: opacity 0.45s ease;
			transition: opacity 0.45s ease;
			-webkit-overflow-scrolling: touch;
			position: absolute;
			top: 0;
			left: 0;
			width: 100%;
			height: 100%;
			padding: 2.75em;
			opacity: 0;
			overflow-y: auto;
		}

			#menu > .inner > ul {
				list-style: none;
				margin: 0 0 1em 0;
				padding: 0;
			}

				#menu > .inner > ul > li {
					padding: 0;
					border-top: solid 1px rgba(255, 255, 255, 0.15);
				}

					#menu > .inner > ul > li a {
						display: block;
						padding: 1em 0;
						line-height: 1.5;
						border: 0;
						color: inherit;
					}

					#menu > .inner > ul > li:first-child {
						border-top: 0;
						margin-top: -1em;
					}

		#menu > .close {
			-moz-transition: opacity 0.45s ease, -moz-transform 0.45s ease;
			-webkit-transition: opacity 0.45s ease, -webkit-transform 0.45s ease;
			-ms-transition: opacity 0.45s ease, -ms-transform 0.45s ease;
			transition: opacity 0.45s ease, transform 0.45s ease;
			-moz-transform: scale(0.25) rotate(180deg);
			-webkit-transform: scale(0.25) rotate(180deg);
			-ms-transform: scale(0.25) rotate(180deg);
			transform: scale(0.25) rotate(180deg);
			-webkit-tap-highlight-color: transparent;
			display: block;
			position: absolute;
			top: 2em;
			left: -6em;
			width: 6em;
			text-indent: 6em;
			height: 3em;
			border: 0;
			font-size: 1em;
			opacity: 0;
			overflow: hidden;
			padding: 0;
			white-space: nowrap;
		}

			#menu > .close:before, #menu > .close:after {
				-moz-transition: opacity 0.2s ease;
				-webkit-transition: opacity 0.2s ease;
				-ms-transition: opacity 0.2s ease;
				transition: opacity 0.2s ease;
				content: '';
				display: block;
				position: absolute;
				top: 0;
				left: 0;
				width: 100%;
				height: 100%;
				background-position: center;
				background-repeat: no-repeat;
				background-size: 2em 2em;
			}

			#menu > .close:before {
				
				opacity: 0;
			}

			#menu > .close:after {}
				opacity: 1;
			}

			#menu > .close:hover:before {
				opacity: 1;
			}

			#menu > .close:hover:after {
				opacity: 0;
			}

		@media screen and (max-width: 736px) {

			#menu {
				-moz-transform: translateX(16.5em);
				-webkit-transform: translateX(16.5em);
				-ms-transform: translateX(16.5em);
				transform: translateX(16.5em);
				width: 16.5em;
			}

				#menu > .inner {
					padding: 2.75em 1.5em;
				}

				#menu > .close {
					top: 0.5em;
					left: -4.25em;
					width: 4.25em;
					text-indent: 4.25em;
				}

					#menu > .close:before, #menu > .close:after {
						background-size: 1.5em 1.5em;
					}

		}

	body.is-menu-visible #wrapper {
		pointer-events: none;
		cursor: default;
		opacity: 0.25;
	}

	body.is-menu-visible #menu {
		-moz-transform: translateX(0);
		-webkit-transform: translateX(0);
		-ms-transform: translateX(0);
		transform: translateX(0);
		visibility: visible;
	}

		body.is-menu-visible #menu > * {
			opacity: 1;
		}

		body.is-menu-visible #menu .close {
			-moz-transform: scale(1.0) rotate(0deg);
			-webkit-transform: scale(1.0) rotate(0deg);
			-ms-transform: scale(1.0) rotate(0deg);
			transform: scale(1.0) rotate(0deg);
			opacity: 1;
		}

/* Main */

	#main {
		padding: 0em 0 6em 0 ;
	}

		@media screen and (max-width: 736px) {

			#main {
				padding: 0em 0 4em 0 ;
			}

		}

/* Footer */

	#footer {
		padding: 5em 0 6em 0 ;
		background-color: #f6f6f6;
	}

		#footer > .inner {
			display: -moz-flex;
			display: -webkit-flex;
			display: -ms-flex;
			display: flex;
			-moz-flex-wrap: wrap;
			-webkit-flex-wrap: wrap;
			-ms-flex-wrap: wrap;
			flex-wrap: wrap;
			-moz-flex-direction: row;
			-webkit-flex-direction: row;
			-ms-flex-direction: row;
			flex-direction: row;
		}

			#footer > .inner > * > :last-child {
				margin-bottom: 0;
			}

			#footer > .inner section:nth-child(1) {
				width: calc(66% - 2.5em);
				margin-right: 2.5em;
			}

			#footer > .inner section:nth-child(2) {
				width: calc(33% - 2.5em);
				margin-left: 2.5em;
			}

			#footer > .inner .copyright {
				width: 100%;
				padding: 0;
				margin-top: 5em;
				list-style: none;
				font-size: 0.8em;
				color: rgba(88, 88, 88, 0.5);
			}

				#footer > .inner .copyright a {
					color: inherit;
				}

				#footer > .inner .copyright li {
					display: inline-block;
					border-left: solid 1px rgba(88, 88, 88, 0.15);
					line-height: 1;
					padding: 0 0 0 1em;
					margin: 0 0 0 1em;
				}

					#footer > .inner .copyright li:first-child {
						border-left: 0;
						padding-left: 0;
						margin-left: 0;
					}

		@media screen and (max-width: 1280px) {

			#footer {
				padding: 5em 0 3em 0 ;
			}

				#footer > .inner section:nth-child(1) {
					width: calc(66% - 1.25em);
					margin-right: 1.25em;
				}

				#footer > .inner section:nth-child(2) {
					width: calc(33% - 1.25em);
					margin-left: 1.25em;
				}

		}

		@media screen and (max-width: 980px) {

			#footer > .inner section:nth-child(1) {
				width: 66%;
				margin-right: 0;
			}

			#footer > .inner section:nth-child(2) {
				width: calc(33% - 2.5em);
				margin-left: 2.5em;
			}

		}

		@media screen and (max-width: 736px) {

			#footer {
				padding: 3em 0 1em 0 ;
			}

				#footer > .inner {
					-moz-flex-direction: column;
					-webkit-flex-direction: column;
					-ms-flex-direction: column;
					flex-direction: column;
				}

					#footer > .inner section:nth-child(1) {
						width: 100%;
						margin-right: 0;
						margin: 3em 0 0 0;
					}

					#footer > .inner section:nth-child(2) {
						-moz-order: -1;
						-webkit-order: -1;
						-ms-order: -1;
						order: -1;
						width: 100%;
						margin-left: 0;
					}

					#footer > .inner .copyright {
						margin-top: 3em;
					}

		}

		@media screen and (max-width: 480px) {

			#footer > .inner .copyright {
				margin-top: 3em;
			}

				#footer > .inner .copyright li {
					border-left: 0;
					padding-left: 0;
					margin: 0.75em 0 0 0;
					display: block;
					line-height: inherit;
				}

					#footer > .inner .copyright li:first-child {
						margin-top: 0;
					}

		}
            
/* Wrapper */

	#wrapper > * > .inner {
		width: 100%;
		max-width: 68em;
		margin: 0 auto;
		padding: 0 2.5em;
	}

		@media screen and (max-width: 736px) {

			#wrapper > * > .inner {
				padding: 0 1.25em;
			}

		}
        
        </style>
"""
html=f"""
            <header id="header">
						<div class="inner">
								<nav>
									<ul>
									</ul>
								</nav>
						</div>
					</header>
            <div id="main">
						<div class="inner">
							<header>
								<h1 class="tmt">
                                T M T<br>
                                E F F I C I E N C Y&nbsp;&nbsp;&nbsp;&nbsp;S U I T E </h1>
								<p>This suite is a collection of ideas and innovations across TMT, aimed towards making our day-to-day tasks easier. We appreciate the contributions made by everyone in building this and sincerely thank them for their efforts. You can provide your suggestions and feedback at the following link: <a href="https://docs.google.com/forms/d/e/1FAIpQLSeDW6VCIKAGK2D7t5a0SiQusFcSgL8O3ukMaYilMiGTWnkB-w/viewform?usp=sharing">Feedback form</a></p>								
							</header>
							<section class="tiles">
								<article class="style1">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic10.jpg", "rb").read()).decode()}" width="400" alt="Base64 Image"/>
									</span>
									<a href="href='?page=module1">
										<h2>News Consolidation</h2>
										<div class="content">
											<p>Collate news articles from Google in one click</p>
										</div>
									</a>
								</article>
								<article class="style2">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic02.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module2">
										<h2>Data Procurement</h2>
										<div class="content">
											<p>Download and format data from various public sources</p>
										</div>
									</a>
								</article>
								<article class="style3">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic10.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module3">
										<h2>Geocoding</h2>
										<div class="content">
											<p>Convert addresses to coordinates and vice-versa</p>
										</div>
									</a>
								</article>
								<article class="style4">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic02.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module4">
										<h2>Classification</h2>
										<div class="content">
											<p>Classify your data into buckets using ChatGPT</p>
										</div>
									</a>
								</article>
								<article class="style5">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic10.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module5">
										<h2>In-house Tools & GPTs</h2>
										<div class="content">
											<p>Various custom GPTs and excel-based tools built by teams in wider TMT/BCN</p>
										</div>
									</a>
								</article>
								<article class="style6">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic02.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module6">
										<h2>Graph Digitizer</h2>
										<div class="content">
											<p>Gather data from images of line/bar charts</p>
										</div>
									</a>
								</article>
                                <article class="style7">
									<span class="image">
										<img src="data:image/jpeg;base64,{base64.b64encode(open("images/pic10.jpg", "rb").read()).decode()} alt="" />
									</span>
									<a href="href='?page=module7">
										<h2>LUMI tab Copy</h2>
										<div class="content">
											<p>Gather data from images of line/bar charts</p>
										</div>
									</a>
								</article>
							</section>
						</div>
					</div>
			</div>
"""
html_1=f"""<h1 class="tmt">
                                N E W S<br>
                                C O N S O L I D A T I O N
                                </h1> 
                            """
html_2=f"""<h1 class="tmt">
                                D A T A<br>
                                P R O C U R E M E N T
                                </h1> 
                            """
html_3=f"""<h1 class="tmt">G E C O D I N G</h1> 
                            """
html_4=f"""<h1 class="tmt">
                                C L A S S I F I C A T I O N
                                </h1> 
                            """
html_5=f"""<h1 class="tmt">
                                I N - H O U S E<br>
                                T O O L S  /  G P T s
                                </h1> 
                            """
html_6=f"""<h1 class="tmt">
                                G R A P H<br>
                                D I G I T I Z E R
                                </h1> 
                            """
html_7=f"""<h1 class="tmt">
                                L U M I<br>
                                T A B  C O P Y
                                </h1> 
                            """
g=0
if route == "home":
    st.set_page_config(
    page_title="TMT Efficiency Suite", layout="wide"
)
    st.markdown(html,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)

    st.stop()
    def display_shapefile_data(shp_path, shx_path=None, dbf_path=None, cpg_path=None, prj_path=None):
            try:
                # Load shapefile
                gdf = gpd.read_file(shp_path)

                # Display top 10 rows
                # print("Top 10 rows of the shapefile:")
                # print(gdf.head(10))

                st.header("Sample Data")
                if 'df' in locals() and df is not None:
                    st.write(gdf.head(15))

            except Exception as e:
                st.write("Error")

elif route == "module1":
    st.set_page_config(
    page_title="News Consolidation", layout="wide")
    st.markdown(html_1,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 1 - News Consolidation"

elif route == "module2":
    st.set_page_config(
    page_title="Data Procurement", layout="wide")
    st.markdown(html_2,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 2 - Data Procurement"

elif route == "module3":
    st.set_page_config(
    page_title="Geocoding", layout="wide")
    st.markdown(html_3,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 3 - Geocoding"

elif route == "module4":
    st.set_page_config(
    page_title="Classification", layout="wide")
    st.markdown(html_4,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 4 - Classification"

elif route == "module5":
    st.set_page_config(
    page_title="Excel Tools", layout="wide")
    st.markdown(html_5,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 6 - Excel Tools"

elif route == "module6":
    st.set_page_config(
    page_title="Graph Digitizer", layout="wide")
    st.markdown(html_6,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 5 - Graph Digitizer"
elif route == "module7":
    st.set_page_config(
    page_title="LUMI tab Copy", layout="wide")
    st.markdown(html_7,unsafe_allow_html=True)
    st.markdown(css,unsafe_allow_html=True)
    feature_selection0 = "Module 4 - LUMI tab Copy"


if feature_selection0 == "Module 2 - Data Procurement": 

    feature_selection = st.selectbox("Select", ["Road Dataset","Tower Dataset", "Administrative boundaries Dataset","GSM Arena (Mobile Features) Extracter"])
    
    # If "Road Dataset" is selected, show input method selection
    if feature_selection == "Road Dataset":
        st.subheader("Road Data Extractor")
        flag=0
        # 🔹 Sidebar: Choose Input Method (Only shown after feature selection)  
        # st.sidebar.header("Select Input Method")
        input_method = st.sidebar.radio("Choose an input method:", ["Enter Bounding Box","Select U.S. State"])
    
        # 🔹 File Output Fields (Ensures They Are Always Visible)
        st.sidebar.subheader("Save File Options")
        file_name = st.sidebar.text_input("Enter File Name (without extension)", "roads_output")
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())
    
        # 🔹 Ensure progress state persists when zooming the map
        if "progress" not in st.session_state:
            st.session_state["progress"] = 0
    
        # 🔹 Load State Shapefile Automatically (Only if user selects "Select U.S. State")
        STATE_SHP_PATH = r"State_polygon\state.shp"
    
        state_boundary = None

        

    
        if input_method == "Select U.S. State":
            if os.path.exists(STATE_SHP_PATH):
                gdf_states = gpd.read_file(STATE_SHP_PATH)
    
                # Ensure correct coordinate reference system (OSM uses EPSG:4326)
                if gdf_states.crs != "EPSG:4326":
                    gdf_states = gdf_states.to_crs(epsg=4326)
    
                # Detect the correct column for state names
                state_column = None
                for col in gdf_states.columns:
                    if col.lower() in ["state_name", "name"]:
                        state_column = col
                        break
    
                if state_column is None:
                    st.error(f"❌ Could not find a valid state name column. Available columns: {list(gdf_states.columns)}")
                    st.stop()
    
                # 🔹 Ensure State Names Are Strings
                gdf_states[state_column] = gdf_states[state_column].astype(str)
    
                # 🔹 Sidebar: Select a Specific State
                st.sidebar.subheader("Select a State")
                state_name = st.sidebar.selectbox("Choose a State", sorted(gdf_states[state_column].unique()))
    
                # ✅ FIX: Use `union_all()` instead of deprecated `unary_union`
                state_boundary = gdf_states[gdf_states[state_column] == state_name].geometry.union_all()
    
        # 🔹 User Enters Bounding Box (Only if user selects "Enter Bounding Box")
        elif input_method == "Enter Bounding Box":
            st.sidebar.subheader("🛠 Enter Bounding Box Coordinates")
            bbox_left = st.sidebar.number_input("Left (West)", value=-91.177961)
            bbox_top = st.sidebar.number_input("Top (North)", value=36.389317)
            bbox_right = st.sidebar.number_input("Right (East)", value=-88.363670)
            bbox_bottom = st.sidebar.number_input("Bottom (South)", value=33.958642)
    
            # Create a bounding box polygon
            state_boundary = box(bbox_left, bbox_bottom, bbox_right, bbox_top)
            state_name = "Custom Bounding Box"
    
        # 🔹 Display the Selected Area on a Map
        def display_shapefile_data(shp_path, shx_path=None, dbf_path=None, cpg_path=None, prj_path=None):
            try:
                # Load shapefile
                gdf = gpd.read_file(shp_path)

                # Display top 10 rows
                # print("Top 10 rows of the shapefile:")
                # print(gdf.head(10))

                st.header("Sample Data")
                if 'df' in locals() and df is not None:
                    st.write(gdf.head(15))

            except Exception as e:
                st.write("Error")
    
        # 🔹 Function to fetch and save road data
        def fetch_osm_data(state_name, state_boundary):
            g=1
            """Fetches and saves road network data from OSM for a given state."""
            try:
                # Persist progress bar state
                if "progress_bar" not in st.session_state:
                    st.session_state["progress_bar"] = st.progress(st.session_state["progress"])
    
                # Fetch road data using multi-threading
                st.session_state["progress"] = 30
                st.session_state["progress_bar"].progress(st.session_state["progress"])
                with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
                    future = executor.submit(ox.graph_from_polygon, state_boundary, network_type="all", retain_all=True, simplify=True)
                    G = future.result()
    
                st.session_state["progress"] = 70
                st.session_state["progress_bar"].progress(st.session_state["progress"])
    
                # Convert road network to GeoDataFrame
                gdf_edges = ox.graph_to_gdfs(G, nodes=False)
    
                # Save the Shapefile
                shapefile_path = os.path.join(save_path, f"{file_name}.shp")
                gdf_edges.to_file(shapefile_path, driver="ESRI Shapefile")
    
                # Zip only the `.shp` file
                zip_path = os.path.join(save_path, f"{file_name}.zip")
                with zipfile.ZipFile(zip_path, "w") as zipf:
                    zipf.write(shapefile_path, os.path.basename(shapefile_path))
    
                st.session_state["progress"] = 100
                st.session_state["progress_bar"].progress(st.session_state["progress"])
    
                return zip_path  # Return zip file path for download
    
            except Exception as e:
                print(f"Error fetching OSM data for {state_name}: {e}")
                return None
            
        if "fetch_osm_clicked" not in st.session_state:
            st.session_state.fetch_osm_clicked = False

        # Initialize flag

        if state_boundary and st.sidebar.button("Fetch Road Data", key="fetch_osm_button"):
            st.session_state.fetch_osm_clicked = True  # Persist button click in session state
            flag = 1  # Set flag

        # if state_boundary and st.session_state.fetch_osm_clicked:
        #     st.subheader(f"Selected Area: {state_name}")
        #     centroid = state_boundary.centroid
        #     m = folium.Map(location=[centroid.y, centroid.x], zoom_start=6, zoom_control=False, scrollWheelZoom=False, dragging=False)
        #     folium.GeoJson(state_boundary, name="Boundary", style_function=lambda x: {
        #         "color": "blue", "weight": 2, "fillOpacity": 0.2
        #     }).add_to(m)
        #     st_folium(m, width=900, height=500, key="map")
        #     k=1

        if state_boundary and g==1:
            st.write(f"Selected Area: {state_name}")
            st.write("Road data extraction may take ~1 sec/Sq.km")
            centroid = state_boundary.centroid
            m = folium.Map(location=[centroid.y, centroid.x], zoom_start=6, zoom_control=False, scrollWheelZoom=False, dragging=False)

            folium.GeoJson(state_boundary, name="Boundary", style_function=lambda x: {
                "color": "blue", "weight": 2, "fillOpacity": 0.2
            }).add_to(m)
            st_folium(m, width=1100, height=500, key="map")

        if g==0:
            st.write(f"Selected Area: {state_name}")
            st.write("Road data extraction may take ~1 sec/Sq.km")
            centroid = state_boundary.centroid
            m = folium.Map(location=[centroid.y, centroid.x], zoom_start=6)
            folium.GeoJson(state_boundary, name="Boundary", style_function=lambda x: {
                "color": "blue", "weight": 2, "fillOpacity": 0.2
            }).add_to(m)
            st_folium(m, width=1100, height=500, key="map")

        # Fetch OSM Road Data Condition
        if (state_boundary and st.session_state.fetch_osm_clicked) or flag == 1:
            zip_path = fetch_osm_data(state_name, state_boundary)
            
            # Display Download Button for Shapefile
            if zip_path and os.path.exists(zip_path):
                with open(zip_path, "rb") as f:
                    st.download_button("⬇ Download Road Dataset", f, f"{state_name}_roads.zip", "application/zip")
                
                st.success("Road dataset saved and ready for download.")
                # display_shapefile_data(f"{save_path}\{file_name}.shp")
            st.session_state.fetch_osm_clicked = False
            
            # Reset flag after execution
            flag = 0
            g=0

        

    if feature_selection == "Tower Dataset":

        # Define OpenCelliD API keys
        API_KEYS = [
            os.getenv("TOWERID1"), 
            os.getenv("TOWERID2")
        ]

        # Function to get a valid API URL
        def get_valid_url():
            for key in API_KEYS:
                url = f"https://www.opencellid.org/downloads.php?token={key}"
                try:
                    hostname=urlparse(url).hostname
                    site_cert = get_certificate_pem(hostname)
                    combined_cert = combine_cert_with_certifi(site_cert)
                    response = secure_request_with_cert(url, combined_cert)
                    # response = requests.get(url, verify=certifi.where())
                    if response.status_code == 200:
                        return url
                except requests.exceptions.RequestException:
                    continue
            return None

        # Fetch the valid URL
        URL = get_valid_url()
        
        # Streamlit UI
        st.subheader("Tower data Extractor")

        st.sidebar.subheader("Save File Options")
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())

        rel=st.sidebar.slider("Since this is a crowdsourced data, its accuracy depends on the number of samples available for each tower. Select the threshold for minimum # of samples below (Recommended value is 20):", min_value=0, max_value=100, value=20)
        # st.sidebar.text("Relevance percentage indicate total number of measurements assigned to the cell tower (ideal value is 20)")

        # Ensure the save directory exists
        save_dir = os.path.join(save_path)
        os.makedirs(save_dir, exist_ok=True)


        COLUMN_NAMES = [
            "Radio", "MCC", "MNC", "AreaCode", "CellCode", "_skip1",
            "Lat", "Lon", "Range", "Sample", "_skip2",
            "CreationDate", "UpdationDate", "_skip3"
        ]
        
        # Columns to keep
        COLUMNS_TO_KEEP = ["Radio", "MCC", "MNC", "AreaCode", "CellCode", "Lon", "Lat", "Range", "Sample", "CreationDate", "UpdationDate"]

        def fetch_country_links():
            """Fetches country-wise download links from OpenCelliD."""
            if not URL:
                st.error("No valid API key available.")
                return {}

            try:
                hostname=urlparse(URL).hostname
                site_cert = get_certificate_pem(hostname)
                combined_cert = combine_cert_with_certifi(site_cert)
                response = secure_request_with_cert(URL, combined_cert)
                # response = requests.get(URL, verify=certifi.where())
                response.raise_for_status()

                soup = BeautifulSoup(response.text, "html.parser")
                tbody = soup.find("tbody")

                country_links = {}
                if tbody:
                    for tr in tbody.find_all("tr"):
                        tds = tr.find_all("td")
                        if len(tds) >= 3:
                            country_name = tds[0].text.strip()
                            links = [a_tag["href"] for a_tag in tds[2].find_all("a", href=True)]
                            if links:
                                country_links[country_name] = links

                return country_links

            except requests.exceptions.RequestException as e:
                st.error(f"Failed to fetch data: {e}")
                return {}

        def download_file(url):
            try:
                """Download a single file using Selenium, retrying with different API keys if needed."""
                download_folder = os.getcwd()
                chrome_options = Options()
                chrome_options.add_argument("--headless")  # Run in headless mode
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--window-size=1920,1080")
                chrome_options.add_experimental_option("prefs", {
                    "download.default_directory": download_folder,
                    "download.prompt_for_download": False,
                    "safebrowsing.enabled": True
                })

                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
            except Exception as e:
                print(f"Error during download: {e}")
                return None

            file_name = os.path.basename(url.split("&file=")[-1])  # Extract file name
            gz_file_path = os.path.join(download_folder, file_name)

            try:
                for key in API_KEYS:  # Iterate through API keys if download fails
                    modified_url = url.replace(f"token={API_KEYS[0]}", f"token={key}")  
                    print(f"Attempting to download with API key: {key}")
                    print(f"Downloading: {modified_url}")

                    driver.get(modified_url)
                    time.sleep(2)  # Allow time for the download to start

                    # Wait for download completion
                    retry_attempts = 5  # Max retries per key
                    while retry_attempts > 0:
                        if os.path.exists(gz_file_path):
                            print(f"Downloaded: {gz_file_path}")
                            driver.quit()
                            return gz_file_path
                        else:
                            time.sleep(3)  # Wait a bit and retry checking
                            retry_attempts -= 1

                    print(f"Download failed using key: {key}, trying next key...")

                print("All API keys failed. Unable to download the file.")
                return None

            except Exception as e:
                print(f"Error during download: {e}")
                return None

            finally:
                driver.quit()


        def extract_gz(gz_file):
            """Extract a single .gz file."""
            if not os.path.exists(gz_file):
                return None

            try:
                csv_file_path = gz_file.replace('.gz', '.csv')
                with gzip.open(gz_file, 'rb') as f_in, open(csv_file_path, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
                os.remove(gz_file)  # Remove the .gz file
                return csv_file_path
            except Exception as e:
                print(f"Failed to extract {gz_file}: {e}")
                return None

        def clean_csv(csv_file):
            """Clean a single CSV file."""
            try:
                df = pd.read_csv(csv_file, header=None, low_memory=False)
                if df.shape[1] < len(COLUMN_NAMES):
                    return None
                df.columns = COLUMN_NAMES
                df = df[COLUMNS_TO_KEEP]
                df["Sample"] = pd.to_numeric(df["Sample"], errors="coerce")
                cleaned_file = csv_file.replace(".csv", "_cleaned.csv")
                df.to_csv(cleaned_file, index=False)
                os.remove(csv_file)
                return cleaned_file
            except Exception as e:
                print(f"Failed to clean {csv_file}: {e}")
                return None 

        def merge_csv_row_wise(csv_files, output_file):
            """Merge multiple CSV files row-wise."""
            if not csv_files:
                return None

            try:
                df_list = [pd.read_csv(file, low_memory=False) for file in csv_files if file is not None]
                combined_df = pd.concat(df_list, axis=0, ignore_index=True)
                combined_df = combined_df[combined_df['Sample'] >= rel]
                combined_df['CreationDate'] = pd.to_datetime(combined_df['CreationDate'], unit='s')
                combined_df['UpdationDate'] = pd.to_datetime(combined_df['UpdationDate'], unit='s')
                combined_df.to_csv(output_file, index=False)
                for file in csv_files:
                    os.remove(file)

                # Show success message in Streamlit
                st.success(f"Download and processing complete. Data saved as: {output_file}")

                return output_file
            except Exception as e:
                print(f"Failed to merge CSV files: {e}")
                return None

        # Fetch data
        country_links = fetch_country_links()

        # st.title("OpenCelliD Country-wise Downloads")
        st.write("Select a country to download the respective dataset.")

        if country_links:
            country = st.selectbox("Choose a country:", list(country_links.keys()))

            if st.button("Download Files"):
                st.write(f"Downloading files for **{country}**:")
                links = country_links[country]
                with ThreadPoolExecutor() as executor:
                    gz_files = list(executor.map(download_file, links))
                gz_files = [f for f in gz_files if f]

                with ThreadPoolExecutor() as executor:
                    csv_files = list(executor.map(extract_gz, gz_files))
                csv_files = [f for f in csv_files if f]

                with ThreadPoolExecutor() as executor:
                    cleaned_csv_files = list(executor.map(clean_csv, csv_files))
                cleaned_csv_files = [f for f in cleaned_csv_files if f]

                merge_csv_row_wise(cleaned_csv_files, f"{save_path}\{country}_tower_data.csv")

        else:
            st.warning("No data available. Please check your internet connection.")





    if feature_selection == "Administrative boundaries Dataset":
        # URL of page
        url = "https://diva-gis.org/data.html"

        # Send a GET request to fetch the page content (disable SSL verification)

        hostname = "diva-gis.org"
        site_cert = get_certificate_pem(hostname)
        combined_cert = combine_cert_with_certifi(site_cert)
        response = secure_request_with_cert(url, combined_cert)

        # pem_path = get_certificate_pem(hostname)
        # response = secure_request_with_cert(url, pem_path)
        print(response)
        # response = requests.get(url, verify=certifi.where())

        # Force UTF-8 encoding
        response.encoding = 'utf-8'

        def extract_zip(zip_path, extract_path):
            """
            Extracts a ZIP file to the specified directory.

            :param zip_path: Path to the ZIP file.
            :param extract_path: Directory where the contents should be extracted.
            """
            try:
                # Ensure the extract path exists
                os.makedirs(extract_path, exist_ok=True)

                # Open the zip file and extract its contents
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)
                
                print(f"Extraction complete! Files are extracted to: {extract_path}")
            except Exception as e:
                print(f"Error extracting ZIP file: {e}")

        # Example usage
        # extract_zip("path/to/your.zip", "path/to/extract/directory")


        def source_link():
            if response.status_code == 200:
                # Parse the HTML content using BeautifulSoup
                soup = BeautifulSoup(response.text, 'html.parser')

                # Find all <option> elements within the select dropdown
                select = soup.find('select', id='cnt')
                # country_p = soup.find('p', string=lambda text: text and 'Country' in text)
                # print("Paragraph:", country_p.text.strip())

                # Get all countries from the <option> tags
                options = select.find_all('option')
                # country_p = soup.find("p", string=lambda text: text and "Country" in text)

                # options = country_p.find_all("option")

                # print(options)

                # Extract and generate country download links
                country_data = []
                for option in options:
                    value = option.get("value")
                    name = option.text.strip()
                    if value:
                        # Extract only the part before "_"
                        short_value = value.split("_")[0]
                        # Construct the download link
                        download_link = f"https://geodata.ucdavis.edu/diva/adm/{short_value}_adm.zip"
                        country_data.append((name, download_link))

                return(country_data)
            
                    

            else:
                print("Failed to retrieve the page. Status code:", response.status_code)
                return None
            
        def download_country_data(download_link,country_name, save_path):
            """
            Download the administrative boundary file for the given country and save it to the specified path.
            """
            
            if not download_link:
                print(f"No download link found for {country_name}.")
                return
            # hostname=urlparse(download_link).hostname
            # site_cert = get_certificate_pem(hostname)
            # combined_cert = combine_cert_with_certifi(site_cert)
            # response = requests.get(url, verify=combined_cert,stream=True)
            # response = secure_request_with_cert(download_link, combined_cert)
            response = requests.get(download_link, stream=True, verify=certifi.where())  
            if response.status_code == 200:
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                
                with open(save_path, 'wb') as file:
                    for chunk in response.iter_content(1024):
                        file.write(chunk)
                print(f"File downloaded successfully: {save_path}")
            else:
                print(f"Failed to download file. Status code: {response.status_code}")



        st.subheader("Administrative data Extractor")

        

        st.sidebar.subheader("Save File Options")
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())

        # Ensure the save directory exists
        save_dir = os.path.join(save_path)
        os.makedirs(save_dir, exist_ok=True)

        # st.title("Administrative Country-wise Downloads")
        st.write("Select a country to download the respective dataset.")

        country_source_links = source_link()

        if country_source_links:
            # Convert list of tuples or dictionaries to a dictionary
            country_dict = {country: links for country, links in country_source_links}

            # Streamlit selectbox for country selection
            country = st.selectbox("Choose a country:", list(country_dict.keys()))

            if st.button("Download All Files"):
                st.write(f"Downloading files for **{country}**:")
                links = country_dict.get(country, [])

                if links:
                    download_country_data(links, country, f"{save_path}\Administrative_{country}.zip")
                    extract_zip(f"{save_path}\Administrative_{country}.zip",f"{save_path}")
                    st.success(f"Download and processing complete. Data saved as: {save_path}\Administartive_{country}.zip")
                else:
                    st.error(f"No links available for {country}.")


    if feature_selection == "GSM Arena (Mobile Features) Extracter":
        st.subheader("GSM Arena (Mobile Features) Extractor")
        st.sidebar.header("GSM Arena (Mobile Features) Extractor")
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())
        def extract_li_text_and_links(class_name, start_url):
            headers = {"User-Agent": "Mozilla/5.0"}

            brand_names = []
            brand_links = {}

            current_url = start_url

            while True:
                response = requests.get(current_url, headers=headers)
                soup = BeautifulSoup(response.content, "html.parser")

                # Extract brands from the current page
                div = soup.find("div", class_=class_name)
                if div:
                    ul = div.find("ul")
                    if ul:
                        li_elements = ul.find_all("li")
                        for li in li_elements:
                            a_tag = li.find("a")
                            if a_tag:
                                text = a_tag.get_text(strip=True)
                                href = "https://www.gsmarena.com/" + a_tag["href"] if not a_tag["href"].startswith("http") else a_tag["href"]
                                if text not in brand_links:  # Avoid duplicates
                                    brand_names.append(text)
                                    brand_links[text] = href

                # Find the next page link
                next_link_tag = soup.find("a", {"title": "Next page", "class": "prevnextbutton"}, href=True)
                print(next_link_tag)
                if not next_link_tag or next_link_tag['href'] == "#":
                    break

                current_url = "https://www.gsmarena.com/" + next_link_tag['href']

            return brand_names, brand_links
        
        def extract_th_content_by_id(url):
            headers = {"User-Agent": "Mozilla/5.0"}

            response = requests.get(url, headers=headers)
            soup = BeautifulSoup(response.content, "html.parser")

            th_texts = []

            # Find the element with id="specs-list"
            specs_list = soup.find(id="specs-list")
            if specs_list:
                th_elements = specs_list.find_all("th")
                for th in th_elements:
                    text = th.get_text(strip=True)
                    if text:
                        th_texts.append(text)

            return th_texts
        

        def extract_spec_section_to_df(url: str, section_name: str, phone_name: str, model_name: str) -> pd.DataFrame:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(url, headers=headers)
            soup = BeautifulSoup(response.content, "html.parser")

            specs = soup.find(id="specs-list")
            if not specs:
                return pd.DataFrame()

            data = {
                "Phone Name": phone_name,
                "Model Name": model_name
            }

            tables = specs.find_all("table")

            for table in tables:
                rows = table.find_all("tr")
                if not rows:
                    continue

                th = rows[0].find("th", scope="row")
                if th and th.get_text(strip=True).lower() == section_name.lower():
                    for row in rows:
                        ttl_td = row.find("td", class_="ttl")
                        nfo_td = row.find("td", class_="nfo")
                        if ttl_td and nfo_td:
                            key = ttl_td.get_text(strip=True)
                            value = nfo_td.get_text(separator=" ", strip=True)
                            data[key] = value
                    break

            return pd.DataFrame([data])
        
        phone_name,phone_link=extract_li_text_and_links("brandmenu-v2 light l-box clearfix","https://www.gsmarena.com/")

        # phone_name.append(" All Phone")
        phone_name.sort()
        # print(phone_name)
        selected_Phone = st.multiselect(
            "Select Phone:",
            phone_name
        )

        if st.button("Show Phone Model", key="show_phone_model_btn"):
            st.session_state["show_phone_model"] = True
            phone_model_name = []
            phone_model_link = {}

            for item in selected_Phone:
                if item in phone_link:
                    phone_model_name1, phone_model_link1 = extract_li_text_and_links("makers", phone_link[item])
                    phone_model_name.extend(phone_model_name1)
                    phone_model_link.update(phone_model_link1)

            # Store in session_state to avoid recomputation
            st.session_state["phone_model_name"] = phone_model_name
            st.session_state["phone_model_link"] = phone_model_link

        # Only show the model selector if models are already fetched
        if st.session_state.get("show_phone_model") and "phone_model_name" in st.session_state:
            selected_model = st.multiselect(
                "Phone Model",
                st.session_state["phone_model_name"]
            )

            # When button is clicked, set flag and store the spec options in session_state
            if st.button("Show Specifications", key="specification_btn"):
                st.session_state["show_specification"] = True
                st.session_state["spec_options"] = [
                    'Network', 'Launch', 'Body', 'Display', 'Platform', 'Memory',
                    'Main Camera', 'Selfie camera', 'Sound', 'Comms', 'Features',
                    'Battery', 'Misc'
                ]

            # Display the multiselect only if the button was clicked
            if st.session_state.get("show_specification") and "spec_options" in st.session_state:
                selected_spec = st.multiselect(
                    "Select Specification:",
                    st.session_state["spec_options"]
                )

            if st.button("Extract data", key="extract"):
            # Fake browser headers to avoid blocking
                HEADERS = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 \
                                (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
                    "Accept-Language": "en-US,en;q=0.9",
                }
                st.session_state["Extract data"] = True
                spec_sheets = {spec: [] for spec in selected_spec}

                progress_bar = st.progress(0, text="Extracting data...")

                total_tasks = len(selected_model) * len(selected_spec)
                task_count = 0

                # First loop over phone models
                for model in selected_model:
                    if model not in st.session_state["phone_model_link"]:
                        continue

                    url = st.session_state["phone_model_link"][model]
                    phone_name = url.split("https://www.gsmarena.com/")[1].split("_")[0]

                    for spec in selected_spec:
                        try:
                            # Optional request to warm-up the session and mimic human behavior
                            requests.get(url, headers=HEADERS, timeout=10)
                            time.sleep(random.uniform(1, 5))  # Avoid bot detection

                            # Your extractor should return a DataFrame with one row
                            df_row = extract_spec_section_to_df(url, spec, phone_name, model)

                            if not df_row.empty:
                                spec_sheets[spec].append(df_row)

                        except Exception as e:
                            st.warning(f"Failed: {model} - {spec}: {e}")

                        task_count += 1
                        progress_bar.progress(task_count / total_tasks, text=f"Processing: {model} - {spec}")

                # Save to Excel with separate sheet for each spec
                file_path = f"{save_path}\\phone_specs.xlsx"
                with pd.ExcelWriter(file_path) as writer:
                    for spec_name, df_list in spec_sheets.items():
                        if df_list:  # Only write if there's data
                            df_combined = pd.concat(df_list, ignore_index=True)
                            sheet_name = spec_name[:31]
                            df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

                progress_bar.empty()
                st.success(f"Data extraction complete. File saved as: `{file_path}`")


if feature_selection0 == "Module 3 - Geocoding":
    page = st.selectbox("Select", ["Geocoder", "Reverse Geocoder"])
    # print(page)

    if page in ["Geocoder", "Reverse Geocoder"]:
        # Tools pages

        # Move the logo to the sidebar
        # st.sidebar.image(r"Logo.png", width = 100)
        # st.sidebar.title("TMT Efficiency Suite")

        if page == "Geocoder":
            st.sidebar.header("Geocoder Tool")
            st.write("This tool converts address lines to latitude and longitude. For the tool to work, please ensure you have the columns for address line and zipcode in your dataset. Geocoding may take approx. 2 sec/record")
        elif page == "Reverse Geocoder":
            st.sidebar.header("Reverse Geocoder Tool")
            st.write("This tool converts latitude and longitude to address lines. For the tool to work, please ensure you have the columns for latitude and longitude in your dataset. Reverse Geocoding may take approx. 3 sec/record")

        st.sidebar.header("Input Parameters")

        # Add a radio button to select file type
        input_file_type = st.sidebar.radio("Select File Type", ("Excel", "CSV"))
        if input_file_type=="Excel":
            file = st.sidebar.file_uploader("Upload File", type=["xlsx"])
        else :
            file = st.sidebar.file_uploader("Upload File", type=["csv"])
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())

        start_time = None

        def geocode_addresses(df,address_col,zip_col,num_threads=5):
            """
            Geocode addresses from a CSV or Excel file using multithreading.

            Parameters:
                file_path (str): Path to the input file.
                input_file_type (str): "CSV" or "Excel" (default: "CSV").
                num_threads (int): Number of threads to use (default: 10).

            Returns:
                DataFrame: Rows with NA values in Latitude and Longitude.
            """

            # Read the input file
            # df = pd.read_excel(file_path) if input_file_type.lower() == "excel" else pd.read_csv(file_path)

            # Specify the columns containing address and zip code (Modify as needed)
            # address_col = "UNIQUE"  # Replace with actual column name
            # zip_col = "POSTCODE"  # Replace with actual column name

            start_time = time.time()

            # Initialize geolocator
            geolocator = Nominatim(user_agent="ab-locator", timeout=10)

            # Ensure 'longitude' and 'latitude' columns exist with proper data type
            df["longitude"] = pd.NA  # Using pd.NA for missing values
            df["latitude"] = pd.NA

            # Function to perform geocoding
            def geocode_address(row):
                add_line_1 = row[address_col]
                zip_code = row[zip_col]
                full_address = f"{add_line_1}, {zip_code}"

                try:
                    location = geolocator.geocode(full_address)
                    if location:
                        return row.name, location.longitude, location.latitude
                except Exception:
                    return row.name, pd.NA, pd.NA  # Assigning pd.NA instead of 'NA'

                return row.name, pd.NA, pd.NA

            # Use ThreadPoolExecutor for multithreading
            print("Geocoding in progress using multithreading...")
            results = []
            i=0
            num_row=df.shape[0]
            with ThreadPoolExecutor(max_workers=num_threads) as executor:
                future_to_row = {executor.submit(geocode_address, row): row for _, row in df.iterrows()}

                for future in as_completed(future_to_row):
                    print(f"completed {i} completed {i} completed {i} completed {i} completed {i} completed {i} completed {i}")
                    i=i+1
                    progress_bar.progress((i/num_row)*0.7)
                    results.append(future.result())


            # Update DataFrame with geocoding results
            for index, lon, lat in results:
                df.at[index, "longitude"] = lon
                df.at[index, "latitude"] = lat

            # end_time = time.time()
            # processing_time = end_time - start_time

            # print(f"Processing completed in {processing_time:.2f} seconds using {num_threads} threads.")

            # Return the rows where Latitude and Longitude are missing
            df_na = df[df["longitude"].isna() | df["latitude"].isna()]
            df_valid = df[df["longitude"].notna() & df["latitude"].notna()]

            return df_na, df_valid  # Return DataFrame with missing geolocation values

        # def setup_chrome_options():
        #     """Set up Chrome options for Selenium."""
        #     chrome_options = Options()
        #     # chrome_options.add_argument("--headless")
        #     chrome_options.add_argument("--start-maximized")  # Open browser maximized
        #     # chrome_options = Options()
        #     # chrome_options.add_argument("--headless")
        #     # chrome_options.add_argument("--headless=new")  # Ensure new headless mode
        #     chrome_options.add_argument("--disable-gpu")  # Disable GPU
        #     # chrome_options.add_argument("--disable-software-rasterizer")  # No software-based rasterization
        #     # chrome_options.add_argument("--use-gl=swiftshader")  # Force SwiftShader (software OpenGL)
        #     # chrome_options.add_argument("--disable-gpu-compositing")  # Disable GPU-based compositing
        #     # chrome_options.add_argument("--disable-accelerated-2d-canvas")  # Disable accelerated 2D canvas rendering
        #     # chrome_options.add_argument("--disable-accelerated-video-decode")  # Disable video decoding acceleration
        #     # chrome_options.add_argument("--disable-accelerated-mjpeg-decode")  # Disable MJPEG decoding acceleration
        #     # chrome_options.add_argument("--no-sandbox")  # Run without sandboxing (needed in some environments)
        #     # chrome_options.add_argument("--disable-dev-shm-usage")  # Prevent shared memory issues


        #     return chrome_options
        
        def save_to_csv(directory, file_name, data,timestamp_prefix):
            """
            Saves data to a CSV file with a timestamp prefix. If the file exists, appends the data; otherwise, creates a new file.

            :param directory: str, Directory where the CSV file should be saved.
            :param file_name: str, Base name of the CSV file (without extension).
            :param data: list of dicts, Data to be saved.
            """
            # Get current date in <yyyymmdd> format
            # date_prefix = datetime.now().strftime("%Y%m%d")
            # timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S")
    
             # Construct full file path
            full_file_name = f"{timestamp_prefix}_{file_name}.csv"
            # Construct full file path
            # full_file_name = f"{date_prefix}_{file_name}.csv"
            file_path = os.path.join(directory, full_file_name)
            
            # Convert data to DataFrame
            df = pd.DataFrame(data)

            # Check if file exists to determine append/write mode
            file_exists = os.path.exists(file_path)
            
            # Save to CSV (append if file exists)
            df.to_csv(file_path, mode='a' if file_exists else 'w', header=not file_exists, index=False)
            
            # print(f"Data {'appended to' if file_exists else 'saved to'} {file_path}")


        def fetch_lat_lon(formatted_address, max_retries=1):
            """Fastest version: Fetch latitude and longitude of an address using Selenium and Google Maps."""
            result = {"latitude": pd.NA, "longitude": pd.NA}
            driver=None
            try:
                # Configure Chrome options for maximum speed
                chrome_options = Options()
                chrome_options.add_argument("--headless=new")  # Use new headless mode for faster performance
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                
                # Initialize WebDriver once (outside loop for efficiency)
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                url = f'https://www.google.com/maps/search/{formatted_address}/'
                driver.get(url)
                
                # Wait until the URL updates with coordinates
                WebDriverWait(driver, 7).until(lambda d: "@" in d.current_url and "," in d.current_url)
                updated_url = driver.current_url

                match = re.search(r"@([-0-9.]+),([-0-9.]+)", updated_url)
                if match:
                    result["latitude"], result["longitude"] = map(float, match.groups())
            except Exception as e:
                print(f"Error fetching coordinates for {formatted_address}: {e}")
            if driver:
                driver.quit()
            print(result)
            return(result)

        def geocode_addresses_google(df,address_col,zip_col):
            """Accepts a DataFrame and returns a geocoded DataFrame with latitude and longitude."""
            if address_col not in df.columns or zip_col not in df.columns:
                raise ValueError("DataFrame must contain 'UNIQUE' and 'POSTCODE' columns.")
            
            formatted_addresses = (df[address_col].astype(str) + " " + df[zip_col].astype(str))\
                .str.replace(" ", "+", regex=True)\
                .str.replace(r"\++", "+", regex=True)
            
            # chrome_options = setup_chrome_options()
            print(formatted_addresses)
            results = []
            results1=[]
            # i=0
            # num_row=df.shape[0]
            with ThreadPoolExecutor(max_workers=8) as executor:
                results1 = list(executor.map(lambda addr: fetch_lat_lon(addr), formatted_addresses))
                # print(results1)
                # i=i+8
                # progress_bar.progress((i/num_row)*0.3+0.7)
                for i, result in enumerate(results1, start=1):
                    results.append(result)
                    
                    print(f"Completed {i} addresses Completed {i} addresses Completed {i} addresses Completed {i} addresses Completed {i} addresses")
            
             # Store results
            # for i, address in enumerate(formatted_addresses, start=1):
            #     result = fetch_lat_lon(address,driver)
            #     results.append(result)
            #     print(f"Completed {i} addresses: {result}")  # Progress output


            lat_lon_df = pd.DataFrame(results)

            df = df.reset_index(drop=True)  # Reset index to align with lat_lon_df
            lat_lon_df = lat_lon_df.reset_index(drop=True)  # Reset lat_lon_df index
            print(lat_lon_df)
            if lat_lon_df is None or lat_lon_df.empty or lat_lon_df.isna().all().all():
              return pd.DataFrame() 
            # Assign values to the correct DataFrame columns
            df.loc[:, "latitude"] = lat_lon_df["latitude"]
            df.loc[:, "longitude"] = lat_lon_df["longitude"]

            return df

        if file is not None:
            # Read the uploaded file based on the user's selection
            if input_file_type == "Excel":
                df = pd.read_excel(file)
                print(df)
            else:  # Assuming CSV as the default choice
                df = pd.read_csv(file)
                print(df)

            # Get a list of column headers for user selection
            column_headers = df.columns.tolist()

            if page == "Geocoder":
                # Allow the user to choose the address and zip code columns
                address_col = st.sidebar.selectbox("Choose Address Column", column_headers)
                zip_col = st.sidebar.selectbox("Choose Zip Code Column", column_headers)
            elif page == "Reverse Geocoder":
                # Allow the user to choose the latitude and longitude columns
                lat_col = st.sidebar.selectbox("Choose Latitude Column", column_headers)
                lon_col = st.sidebar.selectbox("Choose Longitude Column", column_headers)

            # Create a container for "Run" and "Stop" buttons using st.columns
            run_button, stop_button = st.sidebar.columns(2)
            run_button = run_button.button("Run")
            stop_button = stop_button.button("Stop")

            if run_button:
                start_time = time.time()
                
                # Initialize geolocator
                geolocator = Nominatim(user_agent="ab-locator")
                text_placeholder = st.empty()

                if page == "Geocoder":
                    # Define function to perform geocoding
                    def geocode_address(row):
                        add_line_1 = row[address_col]
                        zip_code = row[zip_col]
                        add = f"{add_line_1}, {zip_code}"
                        try:
                            answer = geolocator.geocode(add)
                            if answer:
                                return answer.longitude, answer.latitude
                        except:
                            pass
                        return 'NA', 'NA'
                    # st.text("Geocoding in progress...")
                    text_placeholder.text("Geocoding in progress...")
                    progress_bar = st.progress(0)
                    timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S")
                    # Apply geocoding to the DataFrame
                    # df_na, df_valid = geocode_addresses(df,address_col,zip_col)
                    # save_to_csv(save_path,"encoding",df_valid,timestamp_prefix)
                    # print(df_valid)
                    df_na=df  #change kr dena bhai
                    df_valid_google = geocode_addresses_google(df_na,address_col,zip_col)
                    progress_bar.progress(100)
                    save_to_csv(save_path,"encoding",df_valid_google,timestamp_prefix)
                    # df = pd.concat([df_valid, df_valid_google], ignore_index=True)
                   
                elif page == "Reverse Geocoder":
                    progress_bar = st.progress(0)
                    text_placeholder.text("Reverse Geocoding in progress...")
                    # Define function to perform reverse geocoding
                    # progress_bar = st.progress(0)

                    def reverse_geocode(row):
                        lat = row[lat_col]
                        lon = row[lon_col]
                        try:
                            answer = geolocator.reverse((lat, lon))
                            if answer:
                                return pd.Series({"Address": answer.address})
                        except:
                            return pd.Series({"Address": "NA"})


                    # def reverse_geocode(row):
                    #     print(row)
                    #     lat = row[lat_col]
                    #     # num_row=row.shape[0]
                    #     lon = row[lon_col]
                    #     # j=0
                    #     try:
                    #         answer = geolocator.reverse((lat, lon))
                            
                    #         print(answer)
                    #         if answer:
                    #             location_data = {
                    #                 "Address": answer.address,
                    #                 # "Class": answer.raw['class'],
                    #                 # "Type": answer.raw['type'],
                    #                 # "Address_Type": answer.raw['addresstype'],
                    #                 # "Bounding_Box": answer.raw['boundingbox']
                    #             }
                    #             return pd.Series(location_data)
                    #     except:
                    #         # j=j+1
                    #         print("site blocked")
                    #         # print((j/num_row))
                    #         # progress_bar.progress((j/num_row))
                    #         pass
                    #     return pd.Series({"Address": 'NA'})

                    # st.text("Reverse Geocoding in progress...")
                    # Apply reverse geocoding to the DataFrame
                    # text_placeholder.text("Reverse Geocoding in progress...")
                    # reverse_geocoded_data = df.apply(reverse_geocode(), axis=1)
                    # timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S")
                    # df = pd.concat([df, reverse_geocoded_data], axis=1)
                    # save_to_csv(save_path,"encoding",df,timestamp_prefix)
                    text_placeholder.text("Reverse Geocoding in progress...")

                    results = []  # Store results
                    num_rows = len(df)  # Total rows for progress tracking
                    for i, row in df.iterrows():
                        result = reverse_geocode(row)
                        results.append(result)
                        
                        # Update progress bar dynamically
                        progress_bar.progress((i + 1) / num_rows)
                        
                        # Simulating a delay (optional, for visualization)
                        # time.sleep(0.1)

                    reverse_geocoded_data = pd.DataFrame(results)

                    # Concatenate results with original DataFrame
                    df = pd.concat([df, reverse_geocoded_data], axis=1)

                    # Save results to CSV
                    timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S")
                    save_to_csv(save_path, "encoding", df, timestamp_prefix)

                    text_placeholder.text("Reverse Geocoding Completed!")
                    # st.success("Reverse Geocoding Process Completed Successfully!")
            
            if stop_button:
                st.text("Geocoding process stopped.")
                start_time = None

            # Display the entire DataFrame
            st.header("Sample data")
            if 'df' in locals() and df is not None:
                st.write(df.head(15))

            if 'df' in locals() and df is not None:
                # st.markdown("### Download Output")
                # st.markdown("Please select any one output format -")
                # Create a column for the download buttons on the right-hand side
                download_column = st.columns(2)

                # Download the processed data as a .csv file
                # if 'df' in locals() and df is not None:
                #     csv = df.to_csv(index=False)
                #     csv_bytes = io.BytesIO(csv.encode())
                #     download_column[0].download_button("Download as CSV", csv_bytes, key="processed_data_csv", on_click=None,
                #                                     args=None, mime="text/csv")

                # Download the processed data as an .xlsx file
                # if 'df' in locals() and df is not None:
                #     with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
                #         df.to_excel(tmpfile, index=False)
                #         tmpfile.seek(0)
                #         tmpfile_data = tmpfile.read()
                #     download_column[1].download_button("Download as Excel", tmpfile_data, key="processed_data_xlsx", on_click=None,
                #                                     args=None, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if start_time is not None:
                    text_placeholder.text(f"Progress completed!! and saved with file name {timestamp_prefix}_encoding.csv")

                    end_time = time.time()
                    processing_time = end_time - start_time
                    # st.text("completed")
                    # download_column[0].write(f"Processing Time: {processing_time:.2f} seconds")


if feature_selection0 == "Module 1 - News Consolidation":

    # urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    # OpenAI API Key
    API_KEY=os.getenv("GPT_API")
    st.write("""Depending on your query, news fetching may takes approx. 15-20 minutes\n"""
	)
    st.write("Relevance score (out of 10) indicates how well the article matches the given query")
    def generate_google_queries(user_query: str, api_key: str, number) -> list:
        """
        Generate five optimized Google search queries based on the user's input.
        
        :param user_query: The initial query provided by the user.
        :param api_key: OpenAI API key for making requests.
        :return: A list of five refined search queries.
        """
        client = openai.OpenAI(api_key=api_key)  # Initialize OpenAI client

        prompt = (
            f"""You are an advanced search optimization expert. Given the user's query: {user_query}, 
            generate five highly effective Google search prompts that maximize relevance and accuracy, make sure that you don't provide with the specific year in the prompt until the query itself have specific date. Ensure that:

            1.Each search prompt refines and enhances the intent behind the original query.
            2.The queries are diverse, covering different aspects, synonyms, and search operators (such as quotes, site-specific searches, or advanced filters) to improve results.
            3.They consider user intent (informational, transactional, or navigational) to return the most relevant results.
            4.The queries avoid unnecessary words and focus on precision for better search outcomes.

            Return the five optimized search queries in a numbered list format. Make sure that the you provide only queries list format without explanation about any each queries """
        )

        try:
            response = client.chat.completions.create(
                model="gpt-4o",  
                messages=[
                    {"role": "system", "content": "You are an expert in generating prompt in google search engine."},
                    {"role": "user", "content": prompt},
                ],
            )

            output_text = response.choices[0].message.content.strip()
            queries = output_text.split("\n")  # Assuming each query is on a new line

            # Remove any numbering if present and strip extra whitespace
            queries = [q.lstrip("0123456789.- ").strip() for q in queries if q.strip()]

            return queries[:number]  # Return the top five queries

        except Exception as e:
            # print(f"Error generating search queries: {e}")
            return []

    def summarize_text(text, api_key=API_KEY):
        """Summarizes the given text using OpenAI's ChatGPT API."""
        try:
            client = openai.OpenAI(api_key=api_key)
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an AI assistant that summarizes text."},
                    {"role": "user", "content": f"Summarize the following text: {text}"}
                ],
                temperature=0.5,
                max_tokens=200
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            return f"Error: {str(e)}"

    def extract_article_content(url, date, title):
        """Extracts text content from a news article."""
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/110.0.0.0 Safari/537.36"}
        try:
            hostname = extract_hostname(url)
            # Proceed as before
            site_cert = get_certificate_pem(hostname)
            combined_cert = combine_cert_with_certifi(site_cert)
            # response = secure_request_with_cert(url, combined_cert)
            response = requests.get(url,timeout=10 ,verify=combined_cert)
            # print(response)

            # response = requests.get(url, headers=headers, , verify=certifi.where())
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            return {"link": url, "date": date, "content": summarize_text(soup.get_text())}
        except requests.RequestException as e:
            return {"link": url, "date": date, "content": title}

    def open_url_and_print_content(url, date, title):
        """Opens the given URL in a browser, waits for the page to load, and extracts content."""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless=new")  # Use new headless mode for faster performance
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            
            # Initialize WebDriver once (outside loop for efficiency)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except Exception as e:
            return {"link": url, "date": date, "content": title}
        
        try:
            print(f"Opening URL: {url}")
            driver.get(url)
            
            # Wait for possible redirection (max 5 seconds)
            initial_url = url
            for _ in range(5):
                time.sleep(1)
                if driver.current_url != initial_url:
                    break
            
            final_url = driver.current_url
            article_data = extract_article_content(final_url, date, title)
            return article_data
        
        except Exception as e:
            return {"link": url, "date": date, "content": title}
        
        finally:
            driver.quit()

    def fetch_news_urls(query, start_date, end_date, country='US'):
        """Fetches news article URLs based on the given query and date range."""
        gn = GoogleNews(lang='en', country=country)
        search_results = gn.search(query, from_=start_date, to_=end_date)
        urls = [item.get("link", "No link available") for item in search_results.get('entries', [])]
        dates = [item.get('published', 'No Date') for item in search_results.get('entries', [])]
        titles = [item.get("title", "No Title") for item in search_results.get('entries', [])]
        return urls, dates, titles
    
    def extract_relevance_score(response_content):
        """Extract numerical relevance score from the response text."""
        match = re.search(r'Relevance Score: (\d+)', response_content)
        return int(match.group(1)) if match else None  # Return None if no score is found

    def compute_relevance_scores(data_frame, api_key):
        client = openai.OpenAI(api_key=api_key)
        
        # Initialize columns
        # data_frame['relevance_reason'] = None
        data_frame['relevance_score'] = None  

        for index, row in data_frame.iterrows():
            try:
                prompt = f"""
                    You are an expert in analyzing texts for prompt relevance in scoring.

                    TASK: 
                    Evaluate the following article summary for relevance to the prompt: "{row['query']}".

                    GUIDELINES:
                    - Try to be very lenient and not so strict in giving scoring
                    - If the article directly discusses or provides significant insights into the prompt, label it as **"Relevant"**.
                    - If the article is somewhat related but does not provide substantial insights, label it as **"Partially Relevant"**.
                    - If the article does not address the prompt in a meaningful way, label it as **"Irrelevant"**.

                    RESPONSE FORMAT (MUST FOLLOW STRICTLY):
                    - Relevance: [Relevant / Partially Relevant / Irrelevant]
                    - Relevance Score: [A number between 1-10, where 10 is the most relevant]

                    ARTICLE SUMMARY:
                    "{row['Summary']}" """
                response = client.chat.completions.create(
                    model="gpt-4o",  
                    messages=[
                        {"role": "system", "content": "You are an expert at analyzing texts and determining the relevance of the text"},
                        {"role": "user", "content":prompt}
                    ],
                    temperature=0.5,
                    max_tokens=200
                )
                
                response_content = response.choices[0].message.content
                # print(response_content)  # Print for debugging

                # Extract numerical relevance score
                score = extract_relevance_score(response_content)

                # Store data in DataFrame
                # data_frame.at[index, 'relevance_reason'] = response_content  # Full response
                data_frame.at[index, 'relevance_score'] = score  # Only the numerical score
                
            except Exception as e:
                print(f"Failed to compute relevance score for row {index}: {e}")

        return data_frame


    def filter_dataframe_by_date(df, start_date, end_date):
        """
        Filters the DataFrame based on a given start and end date.
        
        Parameters:
        df (pd.DataFrame): Input DataFrame containing a 'Published Date' column.
        start_date (str): Start date in 'YYYY-MM-DD' format.
        end_date (str): End date in 'YYYY-MM-DD' format.
        
        Returns:
        pd.DataFrame: Filtered DataFrame containing only rows within the date range.
        """
        # Ensure 'Published Date' is in string format before conversion
        df = df.copy()
        if not pd.api.types.is_datetime64_any_dtype(df['Published Date']):
            df['Published Date'] = pd.to_datetime(df['Published Date'], format='%d %b %Y', errors='coerce')
        
        # Convert input start and end dates to datetime
        start_date = pd.to_datetime(start_date)
        end_date = pd.to_datetime(end_date)
        
        # Filter the DataFrame
        filtered_df = df[(df['Published Date'] >= start_date) & (df['Published Date'] <= end_date)]
        
        return filtered_df

    def save_to_excel(df, file_path,query,start_date,end_date):
        """Saves a list of dictionaries to an Excel file."""
        try:
            # df = pd.DataFrame(data_list)

            # Ensure correct column names
            # df.columns = ["URL", "Published Date", "Summary"]
            
            # Convert "Published Date" to a readable format if available
            if "Published Date" in df.columns:
                df["Published Date"] = pd.to_datetime(df["Published Date"], errors='coerce').dt.strftime("%d %b %Y")

            # Remove rows where the summary contains "Error"
            # if "Summary" in df.columns:
            #     df = df[~df["Summary"].str[:20].str.contains("Error", na=False)]
            # if "URL" in df.columns:
            #     df = df[~df["URL"].str[:50].str.contains("https://news.google.com/rss", na=False)]
            df=df.drop_duplicates()
            df=compute_relevance_scores(df,API_KEY)

            # Save to Excel
            df.drop(columns=['query'], inplace=True)
            df=filter_dataframe_by_date(df,start_date,end_date)
            df.to_excel(file_path, index=False)
            print(f"Data successfully saved to {file_path}")

        except Exception as e:
            print(f"Error while saving to Excel: {e}")
            text_placeholder.text(f"Error occurs, try re-running 'run.bat', if error still exists feel free to fill the feedback form")


    def final_result(query,start_date,end_date,country,query_num):
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            future = executor.submit(fetch_news_urls, query, start_date, end_date, country)
            article_urls, published_dates, article_title = future.result()
        
        results = []
        # total_articles = len(article_urls)
        
        # Process articles concurrently
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            results = list(executor.map(open_url_and_print_content, article_urls, published_dates, article_title))

        progress_bar.progress(10+16*(query_num))
        return(results)


    query_user = st.sidebar.text_input("Enter the Query")
    start_date =st.sidebar.text_input("Enter the start date (YYYY-MM-DD)")
    end_date = st.sidebar.text_input("Enter the end date (YYYY-MM-DD)")
    country = "US"
    save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())
    # path = r"C:\\Users\\69454\\OneDrive - Bain\\Documents\\Lumi suite\\Final" 
    text_placeholder = st.empty()
    if st.sidebar.button("Fetch News"):
        text_placeholder.text("Fetching News...")
        # options = webdriver.ChromeOptions()
        # options.add_argument("--start-maximized")
        # options.add_argument("--disable-blink-features=AutomationControlled")
        
        # service = Service(ChromeDriverManager().install())
        # driver = webdriver.Chrome(service=service, options=options)
        progress_bar = st.progress(0)
        results_final=[]
        results=[]
        noOfQuery=4
        
        # query_list=generate_google_queries(query_user,API_KEY,noOfQuery)
        # query_list.append(query_user)
        query_list=[query_user]
        # print(query_list)

        progress_bar.progress(10)
        df = pd.DataFrame(results_final)

        query_num=0
        for query in query_list:
            if query.startswith('"') and query.endswith('"'):
                query=query[1:-1]
            # print(query)
            query_num=query_num+1
            results=final_result(query,start_date,end_date,country,query_num)
            # print(results)
            new_df = pd.DataFrame(results)
            new_df['query'] = query
            # print(new_df)
            df = pd.concat([df, new_df])
            
        # Define file path
        df = df.rename(columns={"link": "URL", "date": "Published Date","content":"Summary"})
        # print(df)

        # file_name = ''.join(re.findall(r'[a-zA-Z]', query_user[:]))
        timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S")
    
             # Construct full file path
        full_file_name = f"{timestamp_prefix}_NewsData.xlsx"    
        save_to_excel(df, f"{save_path}\\{full_file_name}",query_user,start_date,end_date)
        progress_bar.progress(100)
        text_placeholder.text(f"Fetching News completed!!, and saved with file name {full_file_name}")
        
        st.stop()


if feature_selection0 == "Module 4 - LUMI tab Copy":
    st.title("LUMI Tab Copy Utility")

    def extract_json_object(input_path):
        with open(input_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
    
        # Step 2: Remove non-JSON lines like "Test 1" (we'll re-add manually)
        lines = content.splitlines()
        json_lines = [line for line in lines if line.strip().startswith("{")]
        cleaned_content = "\n".join(json_lines)
    
        # Step 3: Extract JSON objects
        json_objects = re.findall(r'{[\s\S]+?}(?=\s*{|\Z)', cleaned_content)
    
        if len(json_objects) < 2:
            raise ValueError("Expected at least two JSON objects")
    
        # Step 4: Load and reformat as compact JSON
        json_obj_1 = json.loads(json_objects[0])
        json_obj_2 = json.loads(json_objects[1])
        return [json_obj_1,json_obj_2]
    
    
    def save(output_path,json_obj_1,json_obj_2,dashboard_name):
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"{dashboard_name}\n")
            f.write(json.dumps(json_obj_1, separators=(",", ":")) + "\n")
            f.write(json.dumps(json_obj_2, separators=(",", ":")) + "\n")
        print(f"Output saved to: {output_path} with 'Test 1' as first line.")
    
    
    
    def replace_tab_content(tab_key, new_content, json_obj):
        if "tabs" not in json_obj:
            print("'tabs' not found in JSON.")
            return False
    
        if tab_key not in json_obj["tabs"]:
            print(f" Tab key '{tab_key}' not found.")
            return False
    
        json_obj["tabs"][tab_key] = new_content
        json_obj["tabs"][tab_key]["tabId"]=tab_key
        # json_obj["tabs"][tab_key]["dashboard"]["selectedTabId"]=tab_key
        # print (json_obj["tabs"][tab_key])
        # print(f"Tab '{tab_key}' successfully updated.")
        return json_obj
    
    
    def get_tab_by_name(json_obj, tab_name):
        """
        Returns the tab content from 'tabs' where 'tabName' matches the given name.
    
        Args:
            json_obj (dict): The full JSON object (e.g., json_obj_2).
            tab_name (str): The name of the tab to search for.
    
        Returns:
            dict: The content of the matching tab.
            None: If no matching tab is found.
        """
        tabs = json_obj.get("tabs", {})
    
        for tab_key, tab_content in tabs.items():
            if tab_content.get("tabName") == tab_name:
                return tab_content
    
        print(f"Tab with name '{tab_name}' not found.")
        return None
    
    def combine_all_tab_parameters(json_obj1, json_obj2):
        """
        Combines 'parameters' from all tabs in both json_obj1 and json_obj2 into one dictionary.

        Args:
            json_obj1 (dict): First JSON object with 'tabs'.
            json_obj2 (dict): Second JSON object with 'tabs'.

        Returns:
            dict: Combined parameters structure with merged 'definitions', 'values', 'sets', etc.
        """
        combined_parameters = {
            "definitions": {},
            "values": {},
            "sets": {},
            "usage": {},
            "crossfilterTokens": {}
        }

        def merge_from_tabs(tabs, combined_parameters):
            for tab_key, tab_content in tabs.items():
                parameters = tab_content.get("parameters", {})
                for section in ["definitions", "values", "sets", "usage", "crossfilterTokens"]:
                    if section in parameters:
                        for key, val in parameters[section].items():
                            if key not in combined_parameters[section]:
                                combined_parameters[section][key] = val
                            else:
                                # Optional: handle or log duplicates
                                pass

        merge_from_tabs(json_obj1.get("tabs", {}), combined_parameters)
        merge_from_tabs(json_obj2.get("tabs", {}), combined_parameters)
        return combined_parameters
    

    def remove_identical_values(combined_parameters):
        """
        Removes identical key-value pairs across all sections of the combined_parameters dictionary.
        
        Args:
            combined_parameters (dict): Dictionary with sections like 'definitions', 'values', etc.
        
        Returns:
            dict: Cleaned dictionary with only unique key-value pairs across all sections.
        """
        seen = {}
        cleaned_parameters = {
            "definitions": {},
            "values": {},
            "sets": {},
            "usage": {},
            "crossfilterTokens": {}
        }

        for section, items in combined_parameters.items():
            for key, val in items.items():
                if (key, val) not in seen.values():
                    cleaned_parameters[section][key] = val
                    seen[(section, key)] = (key, val)  # track by both key and value
                else:
                    # Identical key-value pair exists in another section
                    pass  # Optional: log or handle duplicates

        return cleaned_parameters
    
    
    def merge_table_strings(json_obj1, json_obj2):
        table1 = json_obj1.get("table", "")
        table2 = json_obj2.get("table", "")
    
        # Split into list and remove duplicates
        table_set = set(filter(None, table1.split(",") + table2.split(",")))
    
        # Join back to a single string
        combined_table = ",".join(sorted(table_set))
    
        # Return new JSON based on first one (or customize as needed)
        merged_json = dict(json_obj1)  # copy json_obj1
        merged_json["table"] = combined_table
        return merged_json
    

    def update_dashboard_titles(data, new_title):
        tabs = data.get("tabs", {})
        
        for tab_key, tab_value in tabs.items():
            dashboard = tab_value.get("dashboard")
            if dashboard and "title" in dashboard:
                dashboard["title"] = new_title

        return data

    json_paths = []
    path_tab_dict = {}
    dash_1_input_path = st.text_input(r"Enter the path of the JSON file where the tab should be copied (e.g., C:\user\Dash.json).")
    num_dashboards = st.number_input("Enter the number of dashboards you want to copy from.", min_value=1, step=1)

    for i in range(num_dashboards):
        dash_2_input_path = st.text_input(rf"Enter JSON file path for Dashboard {i + 1} (Ex-> C:\user\Dash.json)", key=f"path_{i}")

        if dash_2_input_path:
            json_paths.append(dash_2_input_path)

            # Use button with unique key
            if st.button("Show tabs", key=f"btn_show_tabs_{i}"):
                # Use a different key for session state (not same as widget key)
                st.session_state[f"show_tabs_flag_{i}"] = True
                json_dash_2_obj_1, json_dash_2_obj_2 = extract_json_object(dash_2_input_path)
                tabs = json_dash_2_obj_2.get("tabs", {})
                st.session_state[f"tab_names_{i}"] = [tab_data.get("tabName", "N/A") for tab_data in tabs.values()]

            # Display multiselect if flag is set
            if st.session_state.get(f"show_tabs_flag_{i}", False):
                predefined_list = st.session_state.get(f"tab_names_{i}", [])
                selected_options = st.multiselect("Select one or more options", predefined_list, key=f"multiselect_{i}")

                # Add to dictionary
                path_tab_dict[dash_2_input_path] = selected_options

    print(path_tab_dict)
    
    output_path = st.text_input(r"Enter the path where the combined JSON file should be saved (e.g., C:\user\Combine.json).")
    generate_clicked = st.button("Generate Dashboard")
    if generate_clicked:
        json_dash_1_obj_1,json_dash_1_obj_2=extract_json_object(dash_1_input_path)
        tabs_dash_1 = json_dash_1_obj_2.get("tabs", {})
        blank_tab_keys = [tab_key for tab_key, tab_data in tabs_dash_1.items() if tab_data.get("tabName", "N/A") == "N/A"]
        i=0
        for key, value in path_tab_dict.items():    
            json_dash_2_obj_1,json_dash_2_obj_2=extract_json_object(key)
            
            for tabs_name in value:
                content=get_tab_by_name(json_dash_2_obj_2,tabs_name)
                replace_content=replace_tab_content(blank_tab_keys[i],content,json_dash_1_obj_2)
                replace_content["parameters"]=combine_all_tab_parameters(replace_content,json_dash_2_obj_2)
                # replace_content["parameters"]=remove_identical_values(replace_content)
                i=i+1
    
            # print(json_dash_1_obj_1)
            replace_content=update_dashboard_titles(replace_content,"Combined_JSON")
            json_dash_1_obj_1=merge_table_strings(json_dash_1_obj_1,json_dash_2_obj_1)
        save(output_path,json_dash_1_obj_1,replace_content,"Combined_JSON")
        st.success("Process completed successfully!")


if feature_selection0 == "Module 4 - Classification":

    feature_selection = st.selectbox("Select", ["Based on Fuzzy Match","Based on Taxonomy"])
    if feature_selection=="Based on Fuzzy Match":
        st.header("Fuzzy Match between two datasets")
        st.write("This feature tries to logically string match the values in two columns (of different datasets) and give a match score for each combination")
        st.sidebar.header("Input Parameters")

        # File type selection
        input_file_type = st.sidebar.radio("Select File Type", ("Excel", "CSV"))
        SCORER_MAP = {
        "ratio - General use when strings are similar in length and order.": fuzz.ratio,
        "partial_ratio - When matching a shorter string inside a longer one.": fuzz.partial_ratio,
        """token_sort_ratio - When word order varies (e.g., "juice apple" vs. "apple juice").""": fuzz.token_sort_ratio,
        "token_set_ratio - Best for unordered, unbalanced strings.": fuzz.token_set_ratio,
        "partial_token_sort_ratio - Short unordered strings within longer ones.": fuzz.partial_token_sort_ratio,
        "partial_token_set_ratio - Extra robust for incomplete or messy strings.": fuzz.partial_token_set_ratio
        }

        # Unique keys for file uploaders
        if input_file_type == "Excel":
            file1 = st.sidebar.file_uploader("Upload File 1", type=["xlsx"], key="file1_excel")
            file2 = st.sidebar.file_uploader("Upload File 2", type=["xlsx"], key="file2_excel")
        else:
            file1 = st.sidebar.file_uploader("Upload File 1", type=["csv"], key="file1_csv")
            file2 = st.sidebar.file_uploader("Upload File 2", type=["csv"], key="file2_csv")

        # Save path input
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())


        def get_best_match(name, choices, scorer_name):
            scorer = SCORER_MAP.get(scorer_name, fuzz.token_sort_ratio)  # default if invalid
            match, score = process.extractOne(name, choices, scorer=scorer)
            return match, score
        
        def match_with_parent(row,column_name2, list2,selected_scorer):
            adv = str(row[column_name2])
            best, score = get_best_match(adv, list2,selected_scorer)
            return best, score

        if file1 is not None and file2 is not None:
            if input_file_type == "Excel":
                df1 = pd.read_excel(file1)
                df2 = pd.read_excel(file2)
            else:
                df1 = pd.read_csv(file1)
                df2 = pd.read_csv(file2)

            # Show previews
            st.subheader("Preview of File 1")
            st.dataframe(df1.head())

            st.subheader("Preview of File 2")
            st.dataframe(df2.head())

            # Dropdowns for selecting any column from each file
            st.subheader("Select Parameters for Matching")
            selected_column_file1 = st.selectbox("Select column from File 1", df1.columns.tolist(), key="col_file1")
            selected_column_file2 = st.selectbox("Select column from File 2", df2.columns.tolist(), key="col_file2")

            selected_scorer = st.selectbox(
                "Select Fuzzy Matching Scorer",
                options=list(SCORER_MAP.keys()),
                index=2  # Default to "token_sort_ratio"
            )


            threshold_keep = st.slider("Select Match Score Threshold", min_value=0, max_value=100, value=80)

            if st.button("Run Fuzzy Match"):
                df1_list = df1[selected_column_file1].dropna().astype(str).tolist()
                df2_list = df2[selected_column_file2].dropna().astype(str).tolist()
                
                progress_bar = st.progress(0)  # Initialize progress bar
                total_rows = len(df1)
                step = 1 / total_rows if total_rows > 0 else 1

                matched_results = []
                for i, row in df1.iterrows():
                    best, score = match_with_parent(row, selected_column_file1, df2_list,selected_scorer)
                    matched_results.append((best, score))
                    progress_bar.progress(min((i + 1) * step, 1.0))  # Update progress

                matched = pd.DataFrame(matched_results, columns=["MatchedName", "MatchScore"])
                df1 = pd.concat([df1, matched], axis=1)
                df1 = df1[df1["MatchScore"] >= threshold_keep]
                
                safe_scorer = selected_scorer.split('-')[0].replace(" ", "_")
                output_file = f"{save_path}\\Fuzzy_result_{safe_scorer}.xlsx"
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    df1.to_excel(writer, sheet_name="Sheet1_Matched", index=False)

                st.success(f"✔ Done! Results written to {output_file}")


    if feature_selection=="Based on Taxonomy":
        st.header("Classification of rows based on Taxonomy (Example: Telco M&A, Survey Verbatims)")
        st.write("This feature takes in a list of descriptions and tries to classify it according to a given taxonomy. Please note that in case you have multiple columns of taxonomy, you would need to create a single concatenated column and then write its definition. Please refer to the below file for sample input schema.")

        with open(
                r"\\file.DCN.bain.com\PDrive\BCC Case Repository\Final\General\WIP\TMT\2025\TMT Efficiency Suite 2025\Sample_Classification_Data.xlsx",
                "rb") as file:
            st.download_button(
                label="Download Sample Input",
                data=file,
                file_name="Sample_Input.xlsx"
            )

        st.sidebar.header("Input Parameters")
        # input_file_type = st.sidebar.radio("Select File Type", ("Excel", "CSV"))
        file = st.sidebar.file_uploader("Upload file having Taxonomy definitions and classification data", type=["xlsx"], key="file_excel")

        # Save path input
        save_path = st.sidebar.text_input("Enter Folder Path to Save File", os.getcwd())
        if file is not None:
            df_definition=pd.read_excel(file, sheet_name="Taxonomy")
            df_deal=pd.read_excel(file, sheet_name="Classification Data")
            st.subheader("Taxonomy preview")
            st.dataframe(df_definition.head()) 
            st.subheader("Row Values preview")
            st.dataframe(df_deal.head())
            
            def parse_response(response):
                parts=[p.strip() for p in response.split(';')]
                if len(parts)<3:
                    return None,None,None,None
                level3_part=parts[2]
                level3, reason, confidence = None, None, None
                if '|' in level3_part:
                    colon_parts=[p.strip() for p in level3_part.split('|')]
                    if len(colon_parts)==3:
                        level3,reason,confidence=colon_parts
                    elif len(colon_parts)>3:
                        level3 = colon_parts[0]
                        confidence=colon_parts[-1]
                        reason = '|'.join(colon_parts[1:-1]).strip()
                    else:
                        level3, reason = colon_parts[0], colon_parts[1] if len(colon_parts) > 1 else None
                        confidence = None
                else:
                    level3=level3_part.strip()
                    reason = ';'.join(parts[3:]).strip() if len(parts) > 3 else None
                    confidence = None
                
                return(parts[0].strip() if len(parts) > 0 else None, parts[1].strip() if len(parts) > 1 else None, level3, reason, confidence)

			# 	# Handle strict format: <Bucket>:<Reason>:<Confidence>
			# 	if ':' in level3_part:
			# 		colon_parts = [p.strip() for p in level3_part.split(':')]
			# 		if len(colon_parts) == 3:
			# 			# Perfect format
			# 			level3, reason, confidence = colon_parts
			# 		elif len(colon_parts) > 3:
			# 			# Extra colons inside reason → join middle pieces back
			# 			level3 = colon_parts[0]
			# 			confidence = colon_parts[-1]
			# 			reason = ':'.join(colon_parts[1:-1]).strip()
			# 		else:
			# 			# Missing confidence → only bucket and reason
			# 			level3, reason = colon_parts[0], colon_parts[1] if len(colon_parts) > 1 else None
			# 			confidence = None
			# 	else:
			# 		# No colon in level3 part
			# 		level3 = level3_part.strip()
			# 		reason = ';'.join(parts[3:]).strip() if len(parts) > 3 else None
			# 		confidence = None

			# 	return(parts[0].strip() if len(parts) > 0 else None, parts[1].strip() if len(parts) > 1 else None, level3, reason, confidence)
				
            
            if st.button("Run"):
                progress_bar = st.progress(0)
                status_text = st.empty()
                def gpt_response(comment=""):
                    definitions = "\n".join(
                    f"{row['Name of Bucket']}: {row['Description']}"
                    for _, row in df_definition.iterrows()
                    )
                    # print(comment)

                    System_prompt=f"""You are a classification expert. You analyze descriptions and classify each of them into the most precise, mutually exclusive buckets based on the bucket's definitions. Only select the **best single match** per description and provide the reason for your choice.
                                """
                    
                    prompt=f"""
                            Classify the following descriptions into the best fit Bucket by selecting the best-fitting option from the provided taxonomy. \n
                            Also, tell about how much confidence you have in your answer. For confidence follows following rules: \n
                             a) If you are sure about the industry of both companies, as well as the nature of the deal, then tag your answer as 'High'.\n
                             b) If you are sure only about one of these things, tag it as 'Medium'.\n
                             c) If you are not sure about any of these or if the entire deal is a bit vague, tag it as 'Low'.\n

                            Definition of each bucket:{definitions} \n

                            Description: {comment} \n

                            Understand the descriptions and provide the best answer out of this, do not create your answer and also provide the reasoning behind your choice. the format should be in following way.

                            Strict format: <Name of Bucket>|<Reason>|<Confidence>

                """
                    # print(prompt)
                    openai.api_key=os.getenv("GPT_API")
                    response=openai.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role":"system","content":System_prompt},
                            {"role":"user","content":prompt}
                        ]
                    )
                    # type(response)
                    # print()
                    # print(response.choices[0].message.content)
                    return(response.choices[0].message.content)
                

                results = [None] * len(df_deal)

                with ThreadPoolExecutor(max_workers=20) as executor:
                    # Submit jobs
                    futures = {executor.submit(gpt_response, note): i for i, note in enumerate(df_deal['Classification Column'])}
                    for completed_count, future in enumerate(as_completed(futures)):
                        idx = futures[future]
                        try:
                            result = future.result()
                        except Exception as e:
                            result = f"Error: {str(e)}"
                        results[idx] = result

                        # Update progress
                        progress = int((completed_count + 1) / len(df_deal) * 100)
                        progress_bar.progress(progress)
                        status_text.text(f"Processed {completed_count + 1} of {len(df_deal)} rows")

                # print(type(result))
                df_deal['gptResponse']=results
                
				 
                parsed = df_deal['gptResponse'].apply(parse_response)
                print(parsed)
                df_deal['Level_1 Taxonomy'] = parsed.apply(lambda x: x[0])
                df_deal['Level_2 Taxonomy'] = parsed.apply(lambda x: x[1])
                df_deal['Level_3 Taxonomy'] = parsed.apply(lambda x: x[2])
                df_deal['Reason'] = parsed.apply(lambda x: x[3])
                df_deal['Confidence'] = df_deal['gptResponse'].apply(lambda text: text.rsplit('|', 1)[-1] if isinstance(text, str) else None)
                # df_deal['Confidence'] = parsed.apply(lambda x: x[4])
                df_deal['Confidence'] = df_deal['Confidence'].fillna('Low')

                df_deal.drop(columns=['gptResponse'],inplace=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{save_path}\\OUTPUT_Taxonomy_{timestamp}.xlsx"
                st.success(f"""Classification completed! Saved with file name "OUTPUT_Taxonomy_{timestamp}""""")
                # print(df_deal)

                df_deal.to_excel(output_file,index=False,sheet_name="Row_Classification")
                bold_font = Font(bold=True, color="FFFFFF")  # white bold text
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                wb = load_workbook(output_file)
                ws = wb["Row_Classification"]
                ws.sheet_view.showGridLines = False
                # Define a thin border for all sides
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Determine the range of cells with data
                max_row = ws.max_row
                max_col = ws.max_column

                # Apply the border to each cell with data
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        if cell.value is not None:  # optional: only style non-empty cells
                            cell.border = thin_border


                for cell in ws[1]:
                    cell.font = bold_font
                    cell.fill = fill

                wb.save(output_file)

if feature_selection0 == "Module 5 - Graph Digitizer":
    feature_selection = st.selectbox("Select", ["Y-coordinates for all x-axis labels", "X-coordinates for all y-axis labels", "All values in a stacked bar chart (Beta)", "Coordinates for all dots on the chart (Beta)"])

    st.write("This feature will give you the rough coordinates for all elements on the chart")

    st.sidebar.header("Input Parameters")
    image_file = st.sidebar.file_uploader("Upload image with graph", type=["png", "jpg", "jpeg"])

    if image_file is not None:

        image = image_file.read()
        st.image(image, caption="Uploaded Graph", use_container_width=True)

        if st.button("Run"):
            openai.api_key = os.getenv("GPT_API")
            # Detect MIME type from filename
            mime_type, _ = mimetypes.guess_type(image_file.name)
            if mime_type not in ["image/png", "image/jpeg", "image/webp", "image/gif"]:
                st.error(f"Unsupported image type: {mime_type}")
            else:
                # Encode to base64
                base64_image = base64.b64encode(image).decode("utf-8")
                data_url = f"data:{mime_type};base64,{base64_image}"

                if feature_selection == "Y-coordinates for all x-axis labels":
                    response = openai.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "user", "content": [
                                {"type": "text",
                                 "text": "In the given image, for all labels on the x-axis, give the corresponding y-coordinate, as accurate as possible. Put this information in a table, and do not include any ~ symbol."},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]}
                        ])
                elif feature_selection == "X-coordinates for all y-axis labels":
                    response = openai.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "user", "content": [
                                {"type": "text",
                                 "text": "In the given image, for all labels on the y-axis, give the corresponding x-coordinate, as accurate as possible. Put this information in a table, and do not include any ~ symbol."},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]}
                        ])
                elif feature_selection == "All values in a stacked bar chart (Beta)":
                    response = openai.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "user", "content": [
                                {"type": "text",
                                 "text": "Give the difference between end value and start value of all colors in each of the bars in this photo. If there are two colors in a bar, the values should be such that they represent only the individual portion of those colors, not the cumulative height. Put this information in a table, and do not include any ~symbol."},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]}
                        ])
                else:
                    response = openai.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "user", "content": [
                                {"type": "text",
                                 "text": "In the given chart, for all dots on the chart, give the corresponding x and y coordinates, as accurate as possible. Put this information in a table, and do not include any ~symbol."},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]}
                        ])

                st.markdown(response.choices[0].message.content)

if feature_selection0 == "Module 6 - Excel Tools":
    st.markdown("#### 1. Survey Cutting Tool")
    st.markdown(
        """
        <style>
        .small-text p {
            margin: 0;
            line-height: 1.8;
        }
        </style>
        <div class="small-text">
            <p>This is an automated survey cutter built by BCN, which takes in the raw data, questionnaire and answer key as inputs, and generates the cuts basis selected questions.</p>
            <p><b>POCs in TMT:</b>  Arpit Jain, Shray Arora</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.write(" ")
    with open(r"\\file.DCN.bain.com\PDrive\BCC Case Repository\Final\General\WIP\TMT\2025\250604_AutonomousNetworkSurvey\Reference material\HLS_Survey_Automation_VShare.xlsm", "rb") as file:
        st.download_button(
            label="Download Model",
            data=file,
            file_name="Survey_cutting_tool.xlsm"
        )
    st.write(" \n")

    st.markdown("#### 2. Policy Pal")
    st.markdown(
        """
        <style>
        .small-text p {
            margin: 0;
            line-height: 1.8;
        }
        </style>
        <div class="small-text">
            <p>Ask this GPT anything regarding the various policies or portals at Bain.</p>
            <p><b>POCs in TMT:</b>  Prashant Singh, Ayush Mallick</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.write(" ")
    st.markdown('<a href="https://chatgpt.com/g/g-67ab28145a5c819180e31d00ceef3e4d-policy-pal" target="_blank">GPT Link</a>', unsafe_allow_html=True)
    st.write(" \n")

    st.markdown("#### 3. VC Sector Lens")
    st.markdown(
        """
        <style>
        .small-text p {
            margin: 0;
            line-height: 1.8;
        }
        </style>
        <div class="small-text">
            <p>Ask this GPT regarding any sector/industry and it will evaluate that as a venture capital opportunity</p>
            <p><b>POCs in TMT:</b>  Venkatesh Mishra, Biswajeet Rath</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.write(" ")
    st.markdown(
        '<a href="https://chatgpt.com/g/g-67aaf2646a848191b4506dbbace0fc30-vc-sector-lens-v3" target="_blank">GPT Link</a>',
        unsafe_allow_html=True)
    st.write(" \n")

    st.markdown("#### 4. LSEG (Refinitiv) Financials Extractor")
    st.markdown(
        """
        <style>
        .small-text p {
            margin: 0;
            line-height: 1.8;
        }
        </style>
        <div class="small-text">
            <p>Ask this GPT about the financials of any public company, and it will source that information from LSEG</p>
            <p><b>POCs in TMT:</b>  Lakshit Pardeshi, Venkatesh Mishra</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.write(" ")
    st.markdown(
        '<a href="https://chatgpt.com/g/g-67bc5dd699e88191aaa38579da41475f-lseg-financials-v3-1" target="_blank">GPT Link</a>',
        unsafe_allow_html=True)
    st.write(" \n")

    st.markdown("#### 5. ALToSQL")
    st.markdown(
        """
        <style>
        .small-text p {
            margin: 0;
            line-height: 1.8;
        }
        </style>
        <div class="small-text">
            <p>Give this GPT any Alteryx workflow or XML, and it will give step-by-step SQL queries to perform the same transformations</p>
            <p><b>POCs in TMT:</b>  Dhruv D, Akhil Aggarwal</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.write(" ")
    st.markdown(
        '<a href="https://chatgpt.com/g/g-67a483163f1881918801b55cc51c22ed-alteryx2sql" target="_blank">GPT Link</a>',
        unsafe_allow_html=True)
    st.write(" \n")